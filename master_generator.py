"""
Clara Health Report Generator — BOY edition (new 8-page 19x13in book design).

Renders all 8 spreads with data overlay. Page 6 is still background-only
while the spirometry JSON contract is finalized.

Uses the same SVG image-extraction approach as render_test_pdf.py to lay
down each spread's background, then overlays text via ReportLab.
"""

import os
import re
import math
import base64
import json
import requests
from io import BytesIO
from datetime import datetime
from xml.etree import ElementTree as ET

import openpyxl
from PIL import Image
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from svglib.svglib import svg2rlg
from reportlab.graphics import renderPDF

# resvg rasterizes vector-format page SVGs (gradients, paths) that the embedded-
# PNG extractor can't handle. Optional: if unavailable we fall back to extraction.
try:
    import resvg_py
    _HAS_RESVG = True
except Exception:
    _HAS_RESVG = False

# qrcode renders the Page-1 cover QR. Optional: if unavailable the QR is skipped.
try:
    import qrcode
    _HAS_QR = True
except Exception:
    _HAS_QR = False


# 19in x 13in book spread in PostScript points
PAGE_WIDTH = 19 * 72   # 1368 pt
PAGE_HEIGHT = 13 * 72  # 936 pt
PAGE_SIZE = (PAGE_WIDTH, PAGE_HEIGHT)

# Designer's SVG viewBox
SVG_VIEW_W = 5700.0
SVG_VIEW_H = 3900.0
SCALE_X = PAGE_WIDTH / SVG_VIEW_W   # 0.24
SCALE_Y = PAGE_HEIGHT / SVG_VIEW_H  # 0.24

SVG_NS = "http://www.w3.org/2000/svg"
XLINK_NS = "http://www.w3.org/1999/xlink"

# Standard page content rectangle in PDF points (x, y, w, h), measured from the
# two-half layout of the normal pages (left+right page art sits inside this box
# with a ~6% print margin). Used to place single-image page exports — which may
# carry no <use> transform — so they align/centre exactly like the other pages.
CONTENT_RECT = (88.6, 47.0, 1190.8, 842.0)

# Shared status capsules at backgrounds/Boy root (text baked as vector <text>).
DOCTOR_VISIT_SVG = {"YES": "Doctor Visit(Yes).svg", "NO": "Doctor Visit(No).svg"}
OBSERVATION_NORMAL_SVG = "Normal Observation.svg"

# Clara ID bottom-strip positions (editor-marked). The "Clara ID:" prefix is
# baked into the template SVG; we only overlay the ID value at these coords.
CLARA_ID_LEFT_POS = (195.5, 94.0)
CLARA_ID_RIGHT_POS = (790.7, 94.0)


# ---------- Diet ("food to be included") strips ----------
# Each diet asset is a self-contained SVG (vector food labels + photos) with its
# own viewBox; it's rasterized and placed with its BOTTOM-LEFT corner at a fixed
# PDF-point anchor — ONE shared anchor for every LEFT page, ONE for every RIGHT
# page. Drawn at a fixed height (width follows the strip's aspect) so every diet
# strip reads at the same size regardless of its individual viewBox.
LEFT_DIET_ANCHOR = (399.8, 169.3)    # PDF pt — bottom-left of strip on left pages
RIGHT_DIET_ANCHOR = (1010.1, 169.3)  # PDF pt — bottom-left of strip on right pages
DIET_HEIGHT = 64.0                   # pt — rendered strip height (266 units * 0.24)

# Page 2 (BMI, left): diet keyed by WHO BMI category.
BMI_DIET_SVG = {
    "Underweight": "Under-Weight_Diet.svg",
    "Normal": "Normal_Diet.svg",
    "Overweight": "Over_Weight_Diet.svg",
    "Obese": "Obese_Diet.svg",
    "Morbidly Obese": "Extremely_Obese_Diet.svg",
}
# Page 4 (Anemia, left): diet keyed by anemia category (lowercased to match
# anemia_category_from_blood_work output).
ANEMIA_DIET_SVG = {
    "normal ( non anemic)": "Haemoglobin_Normal_Diet.svg",
    "mild anemic": "Haemoglobin_Mild_Anemic_Diet.svg",
    "moderate anemic": "Haemoglobin_Moderate_Anemic_Diet.svg",
    "severe anemic": "Haemoglobin_Severe_Anemic_Diet.svg",
}
# Page 5 (Personal Hygiene, right): diet keyed by the worse hair/nail category.
HYGIENE_DIET_SVG = {
    "Excellent": "Personal_Hygiene_Excellent_Diet.svg",
    "Moderate": "Personal_Hygiene_Moderate_Diet.svg",
    "Needs Attention": "Personal_Hygiene_Need_Attention_Diet.svg",
}
# Single (non-categorized) diet strips.
VITALS_DIET_SVG = "Pulse_Rate_And_Oxymetry_Diet.svg"   # Page 3, right
RESPIRATORY_DIET_SVG = "Respiratory_Diet.svg"          # Page 5, left
ENT_DIET_SVG = "ENT_Diet.svg"                          # Page 7, right


# ---------- Clara color palette (clara_font_reference "Color Palette") ----------
HEADING_BLUE = colors.HexColor("#B2E2F2")   # headings, labels, Normal Range, most titles
PATIENT_TEAL = colors.HexColor("#8EC8D1")   # patient info, hemoglobin value, dental circular
TEXT_WHITE = colors.HexColor("#FFFFFF")     # body text, pill text, sub-descriptions
NEAR_BLACK = colors.HexColor("#080606")     # Clara ID footer, food labels, anemia values


# ---------- SVG background parsing (reused pattern from render_test_pdf.py) ----------

def parse_svg_images(svg_path):
    """Extract embedded PNG images and their placements from a Page SVG.

    Returns: [{'png_bytes': b..., 'x': float, 'y': float, 'w': float, 'h': float}, ...]
    Coordinates are in SVG viewBox units.
    """
    tree = ET.parse(svg_path)
    root = tree.getroot()

    image_defs = {}
    for img in root.iter(f"{{{SVG_NS}}}image"):
        img_id = img.get("id")
        href = img.get(f"{{{XLINK_NS}}}href") or img.get("href") or ""
        if not href.startswith("data:image/png;base64,"):
            continue
        image_defs[img_id] = base64.b64decode(href.split(",", 1)[1])

    placed = []
    for use in root.iter(f"{{{SVG_NS}}}use"):
        href = use.get(f"{{{XLINK_NS}}}href") or use.get("href") or ""
        if not href.startswith("#"):
            continue
        ref_id = href[1:]
        if ref_id not in image_defs:
            continue

        w = float(use.get("width", "0").rstrip("px"))
        h = float(use.get("height", "0").rstrip("px"))
        tx, ty = 0.0, 0.0
        m = re.search(
            r"matrix\(\s*1\s*,\s*0\s*,\s*0\s*,\s*1\s*,\s*([-\d.]+)\s*,\s*([-\d.]+)\s*\)",
            use.get("transform", "")
        )
        if m:
            tx = float(m.group(1))
            ty = float(m.group(2))

        placed.append({
            "png_bytes": image_defs[ref_id],
            "x": tx, "y": ty, "w": w, "h": h,
        })

    return placed


def svg_text_content(svg_path):
    """Return the concatenated <text> content of an SVG (handles tspans), or ''.
    Used to read the answer baked into a capsule asset (e.g. 'No', 'Yes',
    'Normal') so we can re-draw it with an embedded font."""
    try:
        root = ET.parse(svg_path).getroot()
    except (ET.ParseError, FileNotFoundError):
        return ""
    parts = []
    for t in root.iter(f"{{{SVG_NS}}}text"):
        txt = "".join(t.itertext()).strip()
        if txt:
            parts.append(txt)
    return " ".join(parts)


def svg_has_vector_content(svg_path):
    """True if the SVG carries real vector art (<path> elements), meaning the
    embedded-PNG extractor would drop most of the page. The old 'flattened'
    pages are pure full-page <image> placements with zero paths."""
    try:
        root = ET.parse(svg_path).getroot()
    except ET.ParseError:
        return False
    return next(root.iter(f"{{{SVG_NS}}}path"), None) is not None


def rasterize_svg_background(c, svg_path, zoom=1):
    """Rasterize a whole page SVG with resvg (handles gradients + paths + embedded
    images) and place it as the full-page background. zoom=1 renders at the
    viewBox native size (5700px wide ≈ 300 dpi on the 19in page); resvg requires
    an integer zoom."""
    png_bytes = bytes(resvg_py.svg_to_bytes(svg_path=svg_path, zoom=int(zoom)))
    pil_img = Image.open(BytesIO(png_bytes))
    # Flatten RGBA onto white. With an alpha channel ReportLab's mask='auto'
    # auto-crops the transparent border and re-anchors the art to the page
    # corner; a fully opaque image fills the viewBox edge-to-edge as intended.
    if pil_img.mode == "RGBA":
        flat = Image.new("RGB", pil_img.size, (255, 255, 255))
        flat.paste(pil_img, mask=pil_img.split()[3])
        pil_img = flat
    c.drawImage(ImageReader(pil_img), 0, 0, width=PAGE_WIDTH, height=PAGE_HEIGHT,
                preserveAspectRatio=False)


def draw_background(c, svg_path):
    """Draw one Page SVG as the canvas background.

    Old 'flattened' pages embed full-page PNGs — we place those directly (fast,
    lossless). Newer vector-format pages (paths + gradients) can't be handled by
    the extractor, so we rasterize the whole SVG with resvg instead.
    """
    if _HAS_RESVG and svg_has_vector_content(svg_path):
        rasterize_svg_background(c, svg_path)
        return

    placed = parse_svg_images(svg_path)
    if not placed:
        raise RuntimeError(f"No embedded PNGs found in {svg_path}")

    # Single-image export (one flattened PNG for the whole spread). These often
    # omit the <use> transform, so their stored x/y/w/h can't be trusted — place
    # the image in the standard content rectangle so it centres like every other
    # page instead of anchoring to the viewBox corner.
    if len(placed) == 1:
        pil_img = Image.open(BytesIO(placed[0]["png_bytes"]))
        cx, cy, cw, ch = CONTENT_RECT
        c.drawImage(ImageReader(pil_img), cx, cy, width=cw, height=ch,
                    preserveAspectRatio=False, mask="auto")
        return

    for img in placed:
        x_pt = img["x"] * SCALE_X
        y_pt = PAGE_HEIGHT - (img["y"] + img["h"]) * SCALE_Y
        w_pt = img["w"] * SCALE_X
        h_pt = img["h"] * SCALE_Y

        pil_img = Image.open(BytesIO(img["png_bytes"]))
        c.drawImage(ImageReader(pil_img), x_pt, y_pt, width=w_pt, height=h_pt,
                    preserveAspectRatio=False, mask="auto")


# ---------- Production JSON parsing ----------

def _iso_to_ddmmyyyy(iso_str):
    """Convert an ISO datetime string to DD/MM/YYYY; '' on failure."""
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y")
    except (ValueError, AttributeError):
        return ""


# BMI categories use WHO cutoffs to match the printed ranges on the Page 2
# figure box (<18.5 / 18.5-24.9 / 25-29.9 / 30-39.9 / 40+). The returned name
# also keys both the highlight-overlay SVG and the Excel reference row:
#   Underweight, Normal, Overweight, Obese (figure "Obesity"), Morbidly Obese (figure "Extreme Obesity").
def bmi_category_from_value(bmi):
    """Return the WHO BMI category name, or '' if unknown."""
    try:
        v = float(bmi)
    except (ValueError, TypeError):
        return ""
    if v < 18.5:
        return "Underweight"
    if v < 25.0:
        return "Normal"
    if v < 30.0:
        return "Overweight"
    if v < 40.0:
        return "Obese"
    return "Morbidly Obese"


# Maps a BMI category to the highlight-overlay SVG filename in backgrounds/Boy/Page 2.
BMI_CATEGORY_SVG = {
    "Underweight": "Under-Weight.svg",
    "Normal": "Normal.svg",
    "Overweight": "Over-Weight.svg",
    "Obese": "Obese.svg",
    "Morbidly Obese": "Extremely-Obese.svg",
}


# Page 2 right-page summary bar chart. Per the editor markings, the three
# rating-column x positions are fixed; rows step down by 22.7 pt in y starting
# from BMI=552.6 (Hemoglobin sits at 529.9, then 22.7 steps down).
SUMMARY_BAR_LEFT_X = 850.0   # left edge of each gradient bar (line starts here)
SUMMARY_LINE_THICKNESS = 6.5  # pt — matches the bar's visual height
SUMMARY_X = {"poor": 889.0, "fair": 1016.4, "excellent": 1137.5}
SUMMARY_ROW_KEYS = [
    "bmi", "hemoglobin", "spirometry", "nervous", "cardio",
    "respiratory", "ent", "eye_vision", "dental",
]
SUMMARY_Y_TOP = 552.6  # BMI row (Hemoglobin row sits at 529.9, per the editor markings)
SUMMARY_Y_GAP = 22.7
SUMMARY_ROW_Y = {k: SUMMARY_Y_TOP - i * SUMMARY_Y_GAP for i, k in enumerate(SUMMARY_ROW_KEYS)}


def bmi_summary_rating(category):
    """Map BMI category -> 'poor' | 'fair' | 'excellent' | None."""
    if category == "Normal":
        return "excellent"
    if category in ("Underweight", "Overweight"):
        return "fair"
    if category in ("Obese", "Morbidly Obese"):
        return "poor"
    return None


# Page 3 left-page (Nervous System) capsule SVG assets per category.
# Categories below match the values in the NERVOUS SYSTEM Excel sheet column G.
NERVOUS_PILL_SVGS = {
    "mental_status": {
        "Alert & Active": "mental/Alert.svg",
        "Drowsy": "mental/Drowsy.svg",
        "Irritable": "mental/Irritable.svg",
    },
    "motor_strength": {
        "Symmetrical/Strong": "motor/Strong.svg",
        "Weakness Noted": "motor/Weakness.svg",
    },
    "reflexes": {
        "Normal/Brisk": "reflexes/Normal.svg",
        "Diminished": "reflexes/Diminished.svg",
        "Asymmetrical": "reflexes/Asymmetrical.svg",
    },
}

# Default "positive" category per sub-parameter (used when JSON has no value yet).
NERVOUS_DEFAULTS = {
    "mental_status": "Alert & Active",
    "motor_strength": "Symmetrical/Strong",
    "reflexes": "Normal/Brisk",
}


def load_nervous_system_reference(xlsx_path):
    """Read the NERVOUS SYSTEM sheet -> {sub_param_key: {category: comment}}.

    Sheet layout: column D names the sub-parameter (Mental Status / Reflexes /
    Motor Strength). G holds the category name, H holds the comment. Each
    sub-parameter spans several rows; D is only set on its first row.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["NERVOUS SYSTEM"]

    # Map Excel sub-parameter labels to our keys
    label_to_key = {
        "mental status": "mental_status",
        "motor strength": "motor_strength",
        "reflexes": "reflexes",
    }

    ref = {"mental_status": {}, "motor_strength": {}, "reflexes": {}}
    current_key = None
    for r in range(3, ws.max_row + 1):
        d_val = ws.cell(r, 4).value  # D column (sub parameter)
        if d_val and str(d_val).strip():
            key = label_to_key.get(str(d_val).strip().lower())
            if key:
                current_key = key
        if not current_key:
            continue
        category = ws.cell(r, 7).value  # G
        comment = ws.cell(r, 8).value   # H
        if category and str(category).strip():
            ref[current_key][str(category).strip()] = (str(comment).strip() if comment else "")
    return ref


# Page 4 left-page (Anemia Screening). The ANEMIA Excel sheet keys rows by
# category in column H; pulls parental guidance (J), fact (K), range (L), foods (M).
def load_anemia_reference(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["ANEMIA"]
    ref = {}
    for r in range(3, ws.max_row + 1):
        cat = ws.cell(r, 8).value  # H
        if not cat or not str(cat).strip():
            continue
        name = str(cat).strip()
        ref[name.lower()] = {
            "category": name,
            "doctor_visit": str(ws.cell(r, 9).value or "").strip(),
            "comment": str(ws.cell(r, 10).value or "").strip(),
            "fact": str(ws.cell(r, 11).value or "").strip(),
            "range": str(ws.cell(r, 12).value or "").strip(),
            "foods": str(ws.cell(r, 13).value or "").strip(),
        }
    return ref


def anemia_category_from_blood_work(blood_work):
    """Derive the anemia category from hemoglobin text + comment in the JSON.

    The API returns descriptive text (no numeric value), so we classify by keywords.
    Categories match the ANEMIA Excel sheet H column.
    """
    text = " ".join(
        str(blood_work.get(k, "") or "") for k in ("hemoglobin_text", "hemoglobin_comment")
    ).lower()
    if "severe" in text:
        return "Severe anemic"
    if "moderate" in text:
        return "Moderate Anemic"
    if "mild" in text:
        return "Mild Anemic"
    if "below" in text or ("anemic" in text and "non" not in text):
        return "Mild Anemic"
    return "Normal ( Non Anemic)"


# Page 4 right-page (Cardiovascular System). Three sub-parameters, each with
# its own category set and pill SVG asset under backgrounds/Boy/Page 4/<sub>/.
CARDIO_PILL_SVGS = {
    "heart_sounds": {
        "S1 S2 Normal": "heart sounds/S1 S2 Normal.svg",
        "Murmur Noted": "heart sounds/Murmur Noted.svg",
        "Tachycardia": "heart sounds/Tachycardia.svg",
    },
    "capillary_refill": {
        "Normal (< 2 sec)": "Cappilary Refill/Normal.svg",
        "Delayed": "Cappilary Refill/Delayed.svg",
    },
    "pulse_quality": {
        "Strong & Regular": "pulse quality/Strong.svg",
        "Weak": "pulse quality/Weak.svg",
        "Irregular": "pulse quality/Irregular.svg",
    },
}

CARDIO_DEFAULTS = {
    "heart_sounds": "S1 S2 Normal",
    "capillary_refill": "Normal (< 2 sec)",
    "pulse_quality": "Strong & Regular",
}


def load_cardio_reference(xlsx_path):
    """Read CARDIOVASCULAR SYSTEM sheet -> {sub_param: {category: comment}}.

    Sheet uses column D for sub-parameter (Heart Sounds / Pulse Quality / Capillary
    Refill), column G for category value, column H for the comment text.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["CARDIOVASCULAR SYSTEM"]
    label_to_key = {
        "heart sounds": "heart_sounds",
        "pulse quality": "pulse_quality",
        "capillary refill": "capillary_refill",
    }
    ref = {"heart_sounds": {}, "pulse_quality": {}, "capillary_refill": {}}
    current = None
    for r in range(3, ws.max_row + 1):
        d_val = ws.cell(r, 4).value
        if d_val and str(d_val).strip():
            key = label_to_key.get(str(d_val).strip().lower())
            if key:
                current = key
        if not current:
            continue
        category = ws.cell(r, 7).value
        comment = ws.cell(r, 8).value
        if category and str(category).strip():
            ref[current][str(category).strip()] = (str(comment).strip() if comment else "")
    return ref


# Page 5 left-page (Respiratory System). Three sub-parameters with capsule SVGs
# under backgrounds/Boy/Page 5/<sub>/.
RESPIRATORY_PILL_SVGS = {
    "breath_sounds": {
        "Normal/Vesicular": "breathing sound/Normal.svg",
        "Wheezing": "breathing sound/Wheezing.svg",
        "Crepitations": "breathing sound/Crepetations.svg",
    },
    "effort": {
        "Easy/Normal": "effort/Normal.svg",
        "Use of Accessory Muscles": "effort/UseOfAccesory.svg",
    },
    "cough": {
        "Absent": "cough/Absent.svg",
        "Productive": "cough/Productive.svg",
        "Dry/Irritative": "cough/Dry-Irritative.svg",
    },
}

RESPIRATORY_DEFAULTS = {
    "breath_sounds": "Normal/Vesicular",
    "effort": "Easy/Normal",
    "cough": "Absent",
}


def load_respiratory_reference(xlsx_path):
    """Read RESPIRATORY SYSTEM sheet -> {sub: {category: comment}}.

    Column D = sub-parameter (Breath Sounds / Effort / Cough), column G = category,
    column H = comment text.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["RESPIRATORY SYSTEM"]
    label_to_key = {
        "breath sounds": "breath_sounds",
        "effort": "effort",
        "cough": "cough",
    }
    ref = {"breath_sounds": {}, "effort": {}, "cough": {}}
    current = None
    for r in range(3, ws.max_row + 1):
        d_val = ws.cell(r, 4).value
        if d_val and str(d_val).strip():
            key = label_to_key.get(str(d_val).strip().lower())
            if key:
                current = key
        if not current:
            continue
        category = ws.cell(r, 7).value
        comment = ws.cell(r, 8).value
        if category and str(category).strip():
            ref[current][str(category).strip()] = (str(comment).strip() if comment else "")
    return ref


# Page 8 (General Examination) — 11 sub-parameters split into two columns:
#   LEFT  : pallor, icterus, cyanosis, clubbing, lympha_denopathy, skin_assessment
#   RIGHT : allergy, bone_and_joint, posture, gait_and_coordination, puberty
# Plus a Doctor Visit Recommended pill at bottom-right.
GENERAL_PILL_SVGS = {
    "pallor": {
        "Absent": "pallor/Absent.svg",
        "Mild": "pallor/Mild.svg",
        "Significant": "pallor/Significant.svg",
    },
    "icterus": {
        "Absent": "icterus/Absent.svg",
        "Visible in Sclera": "icterus/Visible-in-scelera.svg",
        "Visible in Skin": "icterus/Visible-in-skin.svg",
    },
    "cyanosis": {
        "Absent": "cyanosis/Absent.svg",
        "Peripheral (Tips)": "cyanosis/Peripheral.svg",
        "Central": "cyanosis/Central.svg",
    },
    "clubbing": {
        "Absent": "clubbing/Absent.svg",
        "Grade 1-4": "clubbing/Grade-1-4.svg",
    },
    "lympha_denopathy": {
        "Not Palpable": "lympha/Not-palpable.svg",
        "Cervical (Neck)": "lympha/Cervical.svg",
        "Axillary/Inguinal": "lympha/Axillary-inguinal.svg",
    },
    "skin_assessment": {
        "Clear/Healthy": "skin assessment/Clear-Healthy.svg",
        "Dry/Eczema": "skin assessment/Dry-Eczema.svg",
        "Rash/Fungal": "skin assessment/Rash-Fungal.svg",
        "Pigmentation": "skin assessment/Pigmentation.svg",
    },
    "allergy": {
        "None Noted": "allergy/None-Noted.svg",
        "Respiratory": "allergy/Respiratory.svg",
        "Skin/Hives": "allergy/Skin-Hives.svg",
        "Food-related": "allergy/Food-related.svg",
    },
    "bone_and_joint": {
        "Normal Range": "bones and joint/Normal.svg",
        "Tenderness": "bones and joint/Tenderness.svg",
        "Swelling": "bones and joint/Swelling.svg",
        "Deformity": "bones and joint/Deformity.svg",
    },
    "posture": {
        "Excellent": "posture/Excellent.svg",
        "Mild Slumping": "posture/Mild-slumping.svg",
        "Scoliosis Suspicion": "posture/Scolosis-suspicion.svg",
    },
    "gait_and_coordination": {
        "Steady/Normal": "gait and coordination/Normal.svg",
        "Mild Limp": "gait and coordination/Mild-Limp.svg",
        "Coordination Delay": "gait and coordination/Coordination-Delay.svg",
    },
    "puberty": {
        "Pre-pubertal": "puberty/Pre-Pubertal.svg",
        "Early Signs": "puberty/Early-Signs.svg",
        "Age Appropriate": "puberty/Age-Appropriate.svg",
    },
}

# Positive defaults used when the JSON has no value for a field (posture and
# gait_and_coordination are absent from the current API).
GENERAL_DEFAULTS = {
    "pallor": "Absent",
    "icterus": "Absent",
    "cyanosis": "Absent",
    "clubbing": "Absent",
    "lympha_denopathy": "Not Palpable",
    "skin_assessment": "Clear/Healthy",
    "allergy": "None Noted",
    "bone_and_joint": "Normal Range",
    "posture": "Excellent",
    "gait_and_coordination": "Steady/Normal",
    "puberty": "Pre-pubertal",
}


def general_exam_normalize(field, value):
    """Map a raw API value to a known Excel category for the given field.

    Most fields use Excel category strings verbatim (after stripping whitespace).
    Puberty is free-form text ("puberty changes appropriate for age."), so we
    pattern-match it. Falls back to the positive default for unrecognized text.
    """
    if not value:
        return None
    v = str(value).strip()
    if not v:
        return None
    if field == "puberty":
        vl = v.lower()
        if "pre" in vl and "pub" in vl:
            return "Pre-pubertal"
        if "early" in vl:
            return "Early Signs"
        if "age" in vl and "appropriate" in vl:
            return "Age Appropriate"
        if "appropriate" in vl:  # "puberty changes appropriate for age."
            return "Age Appropriate"
        return None
    # Exact-match the Excel categories for this field if any exist.
    options = GENERAL_PILL_SVGS.get(field, {})
    if v in options:
        return v
    # Case/whitespace-insensitive match
    for cat in options:
        if cat.lower() == v.lower():
            return cat
    return None


def doctor_visit_label_from_text(value):
    """Map free-form doctor_visit text to 'YES' or 'NO'."""
    v = (value or "").strip().lower()
    if not v:
        return "NO"
    if v.startswith("no ") or "no doctor visit" in v or "not required" in v or "no immediate" in v:
        return "NO"
    return "YES"


# Page 7 left-page (ENT Examination). Four sub-parameters with capsule SVGs
# under backgrounds/Boy/Page 7/<sub>/. Categories match ENT Examination sheet col G.
ENT_PILL_SVGS = {
    "ears": {
        "Clear & Normal": "ears/Normal.svg",
        "Wax Buildup": "ears/Wax-Buildup.svg",
        "Congested/Red": "ears/Congested.svg",
    },
    "nose": {
        "Patent/Clear": "nose/Clear.svg",
        "Deviated Septum": "nose/Deviated.svg",
        "Allergic Rhinitis Signs": "nose/Allergic.svg",
    },
    "throat": {
        "Healthy Mucosa": "throat/Healthy-Mucosa.svg",
        "Tonsillar Hypertrophy": "throat/Tonsiliar-Hypertrophy.svg",
        "Pharyngitis/Redness": "throat/Redness.svg",
    },
    "hearing_status": {
        "Normal Hearing": "hearing status/Normal.svg",
        "Mild Sensitivity Loss": "hearing status/Mild-Sensitivity-Loss.svg",
        "Moderate Sensitivity Loss": "hearing status/Moderate-Sensitivity.svg",
        "Invalid/Not Done": "hearing status/Invalid-Notdone.svg",
    },
}

ENT_DEFAULTS = {
    "ears": "Clear & Normal",
    "nose": "Patent/Clear",
    "throat": "Healthy Mucosa",
    "hearing_status": "Normal Hearing",
}


# Page 7 LEFT page (Dental Examination). Seven sub-parameter pills arranged
# in two columns + a Dentist Visit Recommended pill. SVG assets live under
# backgrounds/Boy/Page 7/<sub-folder>/<Category>.svg. Category strings match
# the API's `value` field for each sub-parameter (analyticMap.value keys the
# dict below: dental_cavity / nursing_bottle_caries / gum_health /
# other_condition / dental_fluorosis / alignment / oral_hygiene).
DENTAL_PILL_SVGS = {
    "dental_cavity": {
        "None Noted": "dental cavity/None-Noted.svg",
        "Early Decay/Spots": "dental cavity/Early-Decay.svg",
        "Deep Cavities": "dental cavity/Deep-Cavities.svg",
    },
    "nursing_bottle_caries": {
        "Not Present": "nursing bottle caries/Not Present.svg",
        "Early Stage (White spots)": "nursing bottle caries/Early Stage.svg",
        "Advanced Decay": "nursing bottle caries/Decay.svg",
    },
    "gum_health": {
        "Healthy/Pink": "gum health/Healthy.svg",
        "Mild Redness/BOP": "gum health/Mild-Redness.svg",
        "Significant Swelling": "gum health/Swelling.svg",
    },
    "other_condition": {
        "ABSENT": "other conditions/Not-Found.svg",
        "Halitosis (Bad breath)": "other conditions/Haltosis.svg",
        "Mouth Ulcers": "other conditions/Mouth-Ulcers.svg",
        "Tongue Tie": "other conditions/Tongue-Tie.svg",
    },
    "dental_fluorosis": {
        "Absent": "dental fluorosis/Absent.svg",
        "Mild (Faint white lines)": "dental fluorosis/Mild.svg",
        "Moderate/Severe": "dental fluorosis/Moderate.svg",
    },
    "alignment": {
        "Normal/Steady": "alignment/Normal.svg",
        "Crowding": "alignment/Crowding.svg",
        "Malocclusion": "alignment/Malocclusion.svg",
    },
    "oral_hygiene": {
        "Excellent": "oral hygiene/Excellent.svg",
        "Plaque/Tartar Present": "oral hygiene/Plaque-Tartar.svg",
        "Significant Buildup": "oral hygiene/Significant-Buildup.svg",
    },
}

# Pill labels rendered on top of each capsule SVG. Defaults to the filename
# (minus extension, dashes → spaces); override here when the SVG name reads
# awkwardly (e.g. "Not Found" → "Absent" for the positive Other-Conditions case).
DENTAL_PILL_LABELS = {
    ("other_condition", "ABSENT"): "Absent",
}

# Positive defaults per sub-parameter (used when JSON lacks a value).
DENTAL_DEFAULTS = {
    "dental_cavity": "None Noted",
    "nursing_bottle_caries": "Not Present",
    "gum_health": "Healthy/Pink",
    "other_condition": "ABSENT",
    "dental_fluorosis": "Absent",
    "alignment": "Normal/Steady",
    "oral_hygiene": "Excellent",
}


def dental_normalize(field, value):
    """Map a raw API value to a known dental category for the given field.

    API value strings already match the Excel/SVG category names verbatim, so
    a case/whitespace-insensitive match is enough. Returns None for unknowns.
    """
    if not value:
        return None
    v = str(value).strip()
    if not v:
        return None
    options = DENTAL_PILL_SVGS.get(field, {})
    if v in options:
        return v
    for cat in options:
        if cat.lower() == v.lower():
            return cat
    return None


def dental_visit_label_from_text(value):
    """Map dentist-visit free-form text -> 'YES' (visit) or 'NO' (routine)."""
    v = (value or "").strip().lower()
    if not v:
        return "NO"
    if "visit" in v and "dentist" in v:
        return "YES"
    return "NO"


def load_dental_reference(xlsx_path):
    """Read 'DENTAL CHCEKUP' (sic) sheet -> {sub_param_key: {category: comment}}.

    Column D holds the sub-parameter label (DENTAL CAVITIES / GUM HEALTH /
    ALIGNMENT/BITE / NURSING BOTTLE CARIES / DENTAL FLUOROSIS / OTHER CONDITIONS
    / ORAL HYGIENE); G is the category value; H is the comment/observation text.
    Each sub-parameter spans several rows; D is set only on its first row.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["DENTAL CHCEKUP"]
    label_to_key = {
        "dental cavities": "dental_cavity",
        "gum health": "gum_health",
        "alignment/bite": "alignment",
        "nursing bottle caries": "nursing_bottle_caries",
        "dental fluorosis": "dental_fluorosis",
        "other conditions": "other_condition",
        "oral hygiene": "oral_hygiene",
    }
    ref = {k: {} for k in set(label_to_key.values())}
    current = None
    for r in range(3, ws.max_row + 1):
        d_val = ws.cell(r, 4).value
        if d_val and str(d_val).strip():
            key = label_to_key.get(str(d_val).strip().lower())
            if key:
                current = key
        if not current:
            continue
        category = ws.cell(r, 7).value
        comment = ws.cell(r, 8).value
        if category and str(category).strip():
            ref[current][str(category).strip()] = (str(comment).strip() if comment else "")
    return ref


def load_general_examination_reference(xlsx_path):
    """Read the 'General Examination ' sheet -> {field_key: {category: comment}}.

    Column D = sub-parameter label (e.g. 'Pallor', 'ICTERUS', 'BONES & JOINTS').
    Column G = category value, column H = comment/observation text. Each
    sub-parameter spans multiple rows; D is only set on its first row.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["General Examination "]
    label_to_key = {
        "pallor": "pallor",
        "icterus": "icterus",
        "cyanosis": "cyanosis",
        "clubbing": "clubbing",
        "lympha- denopathy": "lympha_denopathy",
        "lympha-denopathy": "lympha_denopathy",
        "lympha denopathy": "lympha_denopathy",
        "skin assessment": "skin_assessment",
        "allergy": "allergy",
        "bones & joints": "bone_and_joint",
        "bones and joints": "bone_and_joint",
        "posture": "posture",
        "gait & co ordination": "gait_and_coordination",
        "gait and coordination": "gait_and_coordination",
        "gait and co ordination": "gait_and_coordination",
        "puberty": "puberty",
    }
    ref = {k: {} for k in set(label_to_key.values())}
    current = None
    for r in range(3, ws.max_row + 1):
        d_val = ws.cell(r, 4).value
        if d_val and str(d_val).strip():
            key = label_to_key.get(str(d_val).strip().lower())
            if key:
                current = key
        if not current:
            continue
        category = ws.cell(r, 7).value
        comment = ws.cell(r, 8).value
        if category and str(category).strip():
            ref[current][str(category).strip()] = (str(comment).strip() if comment else "")
    return ref


def load_ent_reference(xlsx_path):
    """Read the ENT Examination sheet -> {sub_param: {category: {'comment','fact'}}}.

    Column D = sub-parameter (EARS / NOSE / THROAT / HEARING STATUS),
    column G = category value, column H = comment/observation text,
    column K = fact text (only present on Normal Hearing — used for the
    audiometry context box at the bottom of the right page).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["ENT Examination"]
    label_to_key = {
        "ears (hearing & canal)": "ears",
        "ears": "ears",
        "nose (airways)": "nose",
        "nose": "nose",
        "throat (oral cavity)": "throat",
        "throat": "throat",
        "hearing status": "hearing_status",
    }
    ref = {"ears": {}, "nose": {}, "throat": {}, "hearing_status": {}}
    current = None
    for r in range(3, ws.max_row + 1):
        d_val = ws.cell(r, 4).value
        if d_val and str(d_val).strip():
            key = label_to_key.get(str(d_val).strip().lower())
            if key:
                current = key
        if not current:
            continue
        category = ws.cell(r, 7).value
        comment = ws.cell(r, 8).value
        fact = ws.cell(r, 11).value
        if category and str(category).strip():
            ref[current][str(category).strip()] = {
                "comment": (str(comment).strip() if comment else ""),
                "fact": (str(fact).strip() if fact else ""),
            }
    return ref


# Page 5 right-page (Personal Hygiene). Categories on the 3-point scale.
HYGIENE_CATEGORIES = ["Needs Attention", "Moderate", "Excellent"]


def hygiene_category_from_text(value, comment=""):
    """Map API hygiene text -> 'Excellent' | 'Moderate' | 'Needs Attention'."""
    t = " ".join(str(x or "") for x in (value, comment)).lower()
    if not t.strip():
        return "Excellent"
    if "moderate" in t or "moderately" in t or "fair" in t:
        return "Moderate"
    if "unclean" in t or "poorly" in t or "significant wax" in t or "buildup" in t or "needs attention" in t or "poor" in t:
        return "Needs Attention"
    return "Excellent"


def load_hygiene_reference(xlsx_path):
    """Read Personal Hyiegne sheet. Returns {category: comment_text} (shared by hair and nail)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Personal Hyiegne"]
    ref = {}
    for r in range(3, ws.max_row + 1):
        category = ws.cell(r, 7).value  # G
        guidance = ws.cell(r, 10).value  # J - Parental Guidence
        if category and str(category).strip() and guidance and str(guidance).strip():
            cat = str(category).strip()
            if cat not in ref:
                ref[cat] = str(guidance).strip()
    return ref


def load_vitals_reference(xlsx_path):
    """Read the VITALS sheet and return the multi-section guidance text from J4.

    J4 packs three condition sections separated by blank lines:
      NORMAL (...) / HIGH PULSE RATE (...) / LOW OXYGEN (...)
    Returns {"normal": "...", "high_pulse": "...", "low_oxygen": "..."},
    each containing the "Outlook / Actions" lines stripped of the header line.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["VITALS"]
    raw = ws.cell(4, 10).value or ""  # J4
    sections = {"normal": "", "high_pulse": "", "low_oxygen": ""}
    # Split on blank-line paragraph break ("\n\n")
    for paragraph in re.split(r"\n\s*\n", raw):
        lines = [ln for ln in (l.strip() for l in paragraph.splitlines()) if ln]
        if not lines:
            continue
        header = lines[0].lower()
        body = "\n".join(lines[1:]) if len(lines) > 1 else "\n".join(lines)
        if "normal" in header and "high" not in header and "low" not in header:
            sections["normal"] = body
        elif "high pulse" in header or "high pulse rate" in header:
            sections["high_pulse"] = body
        elif "low oxygen" in header:
            sections["low_oxygen"] = body
    return sections


def vitals_pick_section(ref, pulse, oxygen):
    """Choose the matching VITALS conclusion paragraph for the actual readings.

    Priority: LOW OXYGEN > HIGH PULSE > NORMAL. Returns the body text or ''.
    """
    try:
        p = float(pulse) if pulse not in ("", None) else None
    except (TypeError, ValueError):
        p = None
    try:
        o = float(oxygen) if oxygen not in ("", None) else None
    except (TypeError, ValueError):
        o = None
    if o is not None and o < 95:
        return ref.get("low_oxygen", "")
    if p is not None and p > 120:
        return ref.get("high_pulse", "")
    return ref.get("normal", "")


def vitals_observation_label(pulse, oxygen):
    """Short status label for the 'Observation:' field on Page 3 right."""
    try:
        p = float(pulse) if pulse not in ("", None) else None
    except (TypeError, ValueError):
        p = None
    try:
        o = float(oxygen) if oxygen not in ("", None) else None
    except (TypeError, ValueError):
        o = None
    if o is not None and o < 95:
        return "Low Oxygen"
    if p is not None and p > 120:
        return "High Pulse Rate"
    if p is None and o is None:
        return ""
    return "Normal"


def _system_summary_rating(data_dict, defaults, severity_map):
    """Generic 'excellent / fair / poor' rating for a system.

    Compares each sub-parameter's actual category to the positive default.
    Returns 'excellent' when every sub-param matches its default, otherwise
    the worst severity level recorded for any deviation. Unknown deviations
    default to 'fair'. Returns None only if there are no defaults to compare.
    """
    if not defaults:
        return None
    worst = "excellent"
    rank = {"excellent": 0, "fair": 1, "poor": 2}
    for key, default_cat in defaults.items():
        actual = (data_dict or {}).get(key) or default_cat
        if actual == default_cat:
            continue
        level = severity_map.get(actual, "fair")
        if rank[level] > rank[worst]:
            worst = level
    return worst


def nervous_summary_rating(nervous_data):
    severity = {
        "Drowsy": "fair", "Irritable": "fair",
        "Weakness Noted": "poor",
        "Diminished": "fair", "Asymmetrical": "fair",
    }
    return _system_summary_rating(nervous_data, NERVOUS_DEFAULTS, severity)


def cardio_summary_rating(cardio_data):
    severity = {
        "Murmur Noted": "poor", "Tachycardia": "fair",
        "Delayed": "fair",
        "Weak": "fair", "Irregular": "poor",
    }
    return _system_summary_rating(cardio_data, CARDIO_DEFAULTS, severity)


def respiratory_summary_rating(respiratory_data):
    severity = {
        "Wheezing": "fair", "Crepitations": "poor",
        "Use of Accessory Muscles": "poor",
        "Productive": "fair", "Dry/Irritative": "fair",
    }
    return _system_summary_rating(respiratory_data, RESPIRATORY_DEFAULTS, severity)


def ent_summary_rating(ent_data):
    severity = {
        "Wax Buildup": "fair", "Congested/Red": "fair",
        "Deviated Septum": "fair", "Allergic Rhinitis Signs": "fair",
        "Tonsillar Hypertrophy": "fair", "Pharyngitis/Redness": "fair",
        "Mild Sensitivity Loss": "fair", "Moderate Sensitivity Loss": "poor",
        "Invalid/Not Done": "fair",
    }
    return _system_summary_rating(ent_data, ENT_DEFAULTS, severity)


def dental_summary_rating(dental_data):
    severity = {
        "Early Decay/Spots": "fair", "Deep Cavities": "poor",
        "Early Stage (White spots)": "fair", "Advanced Decay": "poor",
        "Mild Redness/BOP": "fair", "Significant Swelling": "poor",
        "Halitosis (Bad breath)": "fair", "Mouth Ulcers": "fair",
        "Tongue Tie": "fair",
        "Mild (Faint white lines)": "fair", "Moderate/Severe": "poor",
        "Crowding": "fair", "Malocclusion": "fair",
        "Plaque/Tartar Present": "fair", "Significant Buildup": "poor",
    }
    return _system_summary_rating(dental_data, DENTAL_DEFAULTS, severity)


def hemoglobin_summary_rating(blood_work):
    """Map hemoglobin info (API text + comment) -> rating. Returns None if absent."""
    text = " ".join(
        str(blood_work.get(k, "") or "") for k in ("hemoglobin_text", "hemoglobin_comment")
    ).lower()
    if not text.strip():
        return None
    if "severe" in text or "moderate" in text:
        return "poor"
    if "mild" in text:
        return "fair"
    if "anemic" in text and "non" not in text:
        return "poor"
    if "below" in text:
        return "poor"
    if "normal" in text or "no signs of anemia" in text or "non-anemic" in text:
        return "excellent"
    return None


def load_bmi_reference(xlsx_path):
    """Read the 'BMI' sheet and return {category_lower: {doctor_visit, conclusion, fact, range, foods}}.

    Column layout (BMI sheet): H=category, I=doctor visit, J=parental guidance
    (conclusion), K=fact, L=range, M=foods.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["BMI "]  # note: the sheet name has a trailing space in the source file

    def clean(v):
        return str(v).strip() if v is not None else ""

    ref = {}
    for r in range(3, ws.max_row + 1):
        category = clean(ws.cell(r, 8).value)  # column H
        if not category:
            continue
        ref[category.lower()] = {
            "category": category,
            "doctor_visit": clean(ws.cell(r, 9).value),    # I
            "conclusion": clean(ws.cell(r, 10).value),     # J
            "fact": clean(ws.cell(r, 11).value),           # K
            "range": clean(ws.cell(r, 12).value),          # L
            "foods": clean(ws.cell(r, 13).value),          # M
        }
    return ref


def parse_production_student(entry):
    """Parse one studentsData[] entry into the flat shape the generator expects.

    Production shape (new API):
      { "student": { ..., "school": {...} }, "campData": [ {parameter, subParameter, value, comment}, ... ] }

    Extracts the student/school block, a screening date from campData[].camp.camp_date,
    and BIOMETRICS & VITALS measurements (height/weight/bmi/pulse/oxymetry/hemoglobin).
    The BMI category is computed from the bmi value.
    """
    student_raw = entry.get("student", {}) or {}
    school = student_raw.get("school", {}) or {}
    camp_data = entry.get("campData", []) or []

    # School logo URL (backend serves it, e.g. http://host/uploads/<school>.png).
    # The exact key isn't fixed yet, so accept the common spellings.
    school_logo_url = (
        school.get("logo") or school.get("logoUrl") or school.get("logo_url")
        or school.get("schoolLogo") or school.get("school_logo")
        or school.get("logoURL") or ""
    )

    gender = (student_raw.get("gender") or "").upper()
    sex = gender[0] if gender else ""

    # Screening date: first camp_date found in campData
    screening_date = ""
    for item in camp_data:
        camp_date = (item.get("camp") or {}).get("camp_date")
        if camp_date:
            screening_date = _iso_to_ddmmyyyy(camp_date)
            break

    student = {
        "name": student_raw.get("name", ""),
        "dob": _iso_to_ddmmyyyy(student_raw.get("date_of_birth", "")),
        "sex": sex,
        "class": student_raw.get("class", ""),
        "section": student_raw.get("section", ""),
        "roll_no": student_raw.get("roll_number", ""),
        "admission_no": student_raw.get("admission_number", ""),
        "clara_id": student_raw.get("claraId", ""),
    }

    # BIOMETRICS & VITALS — keyed by analyticMap.value (note source typos: oximetery, heamoglobin)
    # PERSONAL HYGIENE — keyed similarly (nail_hygeine, hair_hyiegne, ear_hygiene)
    # GENERAL EXAMINATION — pallor / icterus / cyanosis / clubbing / lympha_denopathy /
    #   skin_assessment / allergy / bone_and_joint / puberty / doctor_visit (posture and
    #   gait_and_coordination not yet provided by the API).
    measurements, vitals, blood_work = {}, {}, {}
    hygiene = {}
    general = {}
    dental = {}
    for item in camp_data:
        pname = (item.get("parameter") or {}).get("name", "")
        akey = ((item.get("subParameter") or {}).get("analyticMap") or {}).get("value", "")
        val = item.get("value", "") or ""
        com = item.get("comment", "") or ""
        if pname == "BIOMETRICS & VITALS":
            if akey == "height":
                measurements["height"] = val
            elif akey == "weight":
                measurements["weight"] = val
            elif akey == "bmi":
                measurements["bmi"] = val
            elif akey == "pulse_rate":
                vitals["pulse_rate"] = val
            elif akey == "oximetery":
                vitals["oxymetry"] = val
            elif akey == "heamoglobin":
                blood_work["hemoglobin_text"] = val
                blood_work["hemoglobin_comment"] = com
        elif pname == "PERSONAL HYGIENE":
            if akey == "nail_hygeine":
                hygiene["nail_value"] = val
                hygiene["nail_comment"] = com
            elif akey == "hair_hyiegne":
                hygiene["hair_value"] = val
                hygiene["hair_comment"] = com
            elif akey == "ear_hygiene":
                hygiene["ear_value"] = val
                hygiene["ear_comment"] = com
        elif pname == "GENERAL EXAMINATION":
            # Field name for the SVG/Excel mapping == analyticMap.value, with one
            # rename: doctor_visit text is rendered as a YES/NO pill.
            general[akey] = {"value": val, "comment": com}
        elif pname == "DENTAL CHECKUP":
            # akey ∈ {dental_cavity, nursing_bottle_caries, gum_health,
            # other_condition, dental_fluorosis, alignment, oral_hygiene,
            # dental_visit}. dental_visit drives the YES/NO pill at bottom-left.
            dental[akey] = {"value": val, "comment": com}

    return {
        "camp_name": school.get("schoolName", ""),
        "clara_id_camp": school.get("claraId", ""),
        "screening_date": screening_date,
        "report_year": str(datetime.now().year),
        "student": student,
        "measurements": measurements,
        "vitals": vitals,
        "blood_work": blood_work,
        "hygiene": hygiene,
        "general": general,
        "dental": dental,
        "school_logo_url": school_logo_url,
        "bmi_category": bmi_category_from_value(measurements.get("bmi")),
    }


def load_report_data(json_path, student_index=0):
    """Load JSON and normalize to the generator's flat shape.

    Accepts these shapes:
      1. Production "multiple": {"data": {"studentsData": [ {student, campData}, ... ]}}
         -> parses the entry at student_index.
      2. Production "single":   {"data": {"student": {...}, "campData": [...]}}
         This is what pdf_service.py feeds the generator for ONE student —
         including each element of the /reports/data/multiple array, which the
         service splits itself before calling in. -> parses that one student.
      3. Already-flat test JSON (top-level generator shape, no campData).
    """
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    data = raw.get("data", raw)

    # 1. Production "multiple": an array of student objects under studentsData.
    students = data.get("studentsData") if isinstance(data, dict) else None
    if isinstance(students, list) and students:
        return parse_production_student(students[student_index])

    # 2. Production "single": one student object wrapped under "data". Discriminate
    #    on campData (a list) so we never misfire on already-flat test JSON, which
    #    carries measurements/vitals but no campData.
    if isinstance(data, dict) and isinstance(data.get("campData"), list):
        return parse_production_student(data)

    # 3. Already-flat test JSON.
    return raw


# ---------- Fonts ----------

# ReportLab font names for the Clara typography system (see clara_font_reference).
# These are the names we register the TTFs under and reference everywhere via
# the FONTS style map below. Roboto static cuts are instanced from the Google
# Fonts variable font (wght + wdth axes); Bebas Neue is the OFL static face.
#   Roboto Light(300) / Regular(400) / Medium(500) / Bold(700) / Black(900)
#   Roboto Condensed Regular(400, 75% width) / Bold(700, 75% width)
#   Bebas Neue Regular
# Helvetica (FEV₁ subscript) and Lucida Grande (vision-table arrows) are Page-6
# only; Page 6 is rendered background-only, so Helvetica falls back to the
# ReportLab built-in and Lucida Grande needs no registration.
ROBOTO_FONT_FILES = [
    ("Roboto-Light", "Roboto-Light.ttf"),
    ("Roboto", "Roboto-Regular.ttf"),
    ("Roboto-Medium", "Roboto-Medium.ttf"),
    ("Roboto-Bold", "Roboto-Bold.ttf"),
    ("Roboto-Black", "Roboto-Black.ttf"),
    ("RobotoCondensed", "RobotoCondensed-Regular.ttf"),
    ("RobotoCondensed-Bold", "RobotoCondensed-Bold.ttf"),
    ("BebasNeue", "BebasNeue-Regular.ttf"),
]


def register_fonts(fonts_folder):
    """Register the Clara Roboto/Bebas typography set. Returns True if the core
    Roboto faces registered (the FONTS map falls back to Helvetica otherwise)."""
    if not fonts_folder or not os.path.exists(fonts_folder):
        return False

    registered = set()
    for name, filename in ROBOTO_FONT_FILES:
        path = os.path.join(fonts_folder, filename)
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(name, path))
                registered.add(name)
            except Exception as e:
                print(f"⚠️  Failed to register {filename}: {e}")
        else:
            print(f"⚠️  Missing font file: {path}")

    # Core faces required to consider the custom typography "available".
    core = {"Roboto-Light", "Roboto", "Roboto-Bold", "Roboto-Black", "BebasNeue"}
    return core.issubset(registered)


# ---------- Generator ----------

class ClaraBoyReportGenerator:
    """Generates the 8-spread Boy report (19x13in book format)."""

    def __init__(self, backgrounds_root, fonts_folder=None, xlsx_path=None, is_girl=False):
        # backgrounds_root should point to `backgrounds/Images` (boy + girl share
        # this one folder; only the page-background SVG differs by gender).
        self.backgrounds_root = backgrounds_root
        # When True, page backgrounds use the `PageN_Girl.svg` variant. All capsule
        # / pill overlay assets are shared between boy and girl.
        self.is_girl = is_girl
        self.has_custom_fonts = register_fonts(fonts_folder) if fonts_folder else False
        self.xlsx_path = xlsx_path
        self._bmi_ref = None
        self._vitals_ref = None
        self._nervous_ref = None
        self._anemia_ref = None
        self._cardio_ref = None
        self._respiratory_ref = None
        self._hygiene_ref = None
        self._ent_ref = None
        self._general_ref = None
        self._dental_ref = None

    def bmi_reference(self):
        """Lazy-load and cache the BMI reference sheet."""
        if self._bmi_ref is None and self.xlsx_path and os.path.exists(self.xlsx_path):
            self._bmi_ref = load_bmi_reference(self.xlsx_path)
        return self._bmi_ref or {}

    def vitals_reference(self):
        """Lazy-load and cache the VITALS reference sheet."""
        if self._vitals_ref is None and self.xlsx_path and os.path.exists(self.xlsx_path):
            self._vitals_ref = load_vitals_reference(self.xlsx_path)
        return self._vitals_ref or {}

    def nervous_reference(self):
        """Lazy-load and cache the NERVOUS SYSTEM reference sheet."""
        if self._nervous_ref is None and self.xlsx_path and os.path.exists(self.xlsx_path):
            self._nervous_ref = load_nervous_system_reference(self.xlsx_path)
        return self._nervous_ref or {}

    def anemia_reference(self):
        """Lazy-load and cache the ANEMIA reference sheet."""
        if self._anemia_ref is None and self.xlsx_path and os.path.exists(self.xlsx_path):
            self._anemia_ref = load_anemia_reference(self.xlsx_path)
        return self._anemia_ref or {}

    def cardio_reference(self):
        """Lazy-load and cache the CARDIOVASCULAR SYSTEM reference sheet."""
        if self._cardio_ref is None and self.xlsx_path and os.path.exists(self.xlsx_path):
            self._cardio_ref = load_cardio_reference(self.xlsx_path)
        return self._cardio_ref or {}

    def respiratory_reference(self):
        """Lazy-load and cache the RESPIRATORY SYSTEM reference sheet."""
        if self._respiratory_ref is None and self.xlsx_path and os.path.exists(self.xlsx_path):
            self._respiratory_ref = load_respiratory_reference(self.xlsx_path)
        return self._respiratory_ref or {}

    def hygiene_reference(self):
        """Lazy-load and cache the Personal Hyiegne reference sheet."""
        if self._hygiene_ref is None and self.xlsx_path and os.path.exists(self.xlsx_path):
            self._hygiene_ref = load_hygiene_reference(self.xlsx_path)
        return self._hygiene_ref or {}

    def ent_reference(self):
        """Lazy-load and cache the ENT Examination reference sheet."""
        if self._ent_ref is None and self.xlsx_path and os.path.exists(self.xlsx_path):
            self._ent_ref = load_ent_reference(self.xlsx_path)
        return self._ent_ref or {}

    def general_examination_reference(self):
        """Lazy-load and cache the General Examination reference sheet."""
        if self._general_ref is None and self.xlsx_path and os.path.exists(self.xlsx_path):
            self._general_ref = load_general_examination_reference(self.xlsx_path)
        return self._general_ref or {}

    def dental_reference(self):
        """Lazy-load and cache the DENTAL CHCEKUP reference sheet."""
        if self._dental_ref is None and self.xlsx_path and os.path.exists(self.xlsx_path):
            self._dental_ref = load_dental_reference(self.xlsx_path)
        return self._dental_ref or {}

    def _draw_overlay_svg(self, c, svg_relpath, x, y, w, h):
        """Place the first embedded PNG of a Page-folder SVG asset at (x, y) with size (w, h).

        svg_relpath is relative to backgrounds_root (e.g. 'Page 3/mental/Alert.svg').
        """
        full_path = os.path.join(self.backgrounds_root, svg_relpath)
        placed = parse_svg_images(full_path)
        if not placed:
            return
        pil = Image.open(BytesIO(placed[0]["png_bytes"]))
        c.drawImage(ImageReader(pil), x, y, width=w, height=h,
                    preserveAspectRatio=False, mask="auto")

    def _draw_diet(self, c, svg_relpath, anchor, height=DIET_HEIGHT, zoom=2):
        """Render a diet ("food to be included") strip and place it with its
        BOTTOM-LEFT corner at the PDF-point `anchor`, at a fixed height (width
        follows the strip's aspect ratio).

        The strip carries vector food labels alongside photos, so it's rasterized
        with resvg (the embedded-PNG extractor would drop the labels); falls back
        to the embedded PNG if resvg is unavailable.
        """
        full = os.path.join(self.backgrounds_root, svg_relpath)
        if not os.path.exists(full):
            return
        try:
            vb = (ET.parse(full).getroot().get("viewBox") or "").split()
            vw, vh = float(vb[2]), float(vb[3])
        except Exception:
            return
        if not vw or not vh:
            return
        w_pt = height * (vw / vh)
        ax, ay = anchor
        img = None
        if _HAS_RESVG:
            try:
                img = Image.open(BytesIO(bytes(resvg_py.svg_to_bytes(svg_path=full, zoom=int(zoom)))))
            except Exception as e:
                print(f"⚠️  Diet render failed for {svg_relpath}: {e}")
        if img is None:
            placed = parse_svg_images(full)
            if not placed:
                return
            img = Image.open(BytesIO(placed[0]["png_bytes"]))
        c.drawImage(ImageReader(img), ax, ay, width=w_pt, height=height,
                    preserveAspectRatio=False, mask="auto")

    def _draw_pill_capsule(self, c, svg_relpath, x_left, y_center, w=170, h=46):
        """Place a status-capsule (Doctor Visit Yes/No, Normal Observation) at
        (x_left = pill left edge, y_center = pill vertical centre).

        The capsule asset is a coloured PNG shape with its answer text baked as a
        vector <text> element. resvg only renders that text if the system has the
        SVG's font (Arial) installed — which fails on bare Linux/Docker. So we
        place the capsule PNG shape and re-draw its text with our embedded
        Roboto-Bold instead — identical look, works everywhere.
        """
        full_path = os.path.join(self.backgrounds_root, svg_relpath)
        if not os.path.exists(full_path):
            return
        self._draw_overlay_svg(c, svg_relpath, x_left, y_center - h / 2, w, h)
        text = svg_text_content(full_path)
        if text:
            self._draw_pill_label(c, text, x_left, y_center, w)

    # Semantic style -> Roboto/Bebas face, per clara_font_reference. Each role
    # maps to the exact font the designer specified for that kind of text.
    #   patient   : Bebas Neue — Page-1 patient detail values (Name/DOB/Sex…)
    #   footer    : Roboto Black — repeating Clara ID footer
    #   fact      : Roboto Black — fact / stat callout boxes
    #   title     : Roboto Bold — section/page titles (38pt)
    #   label     : Roboto Bold — section field labels, Normal Range, comments-header
    #   pill      : Roboto Bold — status pills (Normal, Alert, NO, …)
    #   value     : Roboto Light — big numeric values + units
    #   body      : Roboto Light — body paragraphs + sub-descriptions
    #   category  : Roboto Regular — summary category labels, parental guidance
    #   food      : Roboto Medium — food labels (Leafy Veggies, Beets, …)
    #   condensed / condensed_bold : Roboto Condensed — Page-1 prevention tagline
    _STYLE_TO_FACE = {
        "patient": "BebasNeue",
        "footer": "Roboto-Black",
        "fact": "Roboto-Black",
        "title": "Roboto-Bold",
        "label": "Roboto-Bold",
        "pill": "Roboto-Bold",
        "value": "Roboto-Light",
        "body": "Roboto-Light",
        "category": "Roboto",
        "food": "Roboto-Medium",
        "condensed": "RobotoCondensed",
        "condensed_bold": "RobotoCondensed-Bold",
        # legacy aliases used by older call sites
        "regular": "Roboto-Light",
        "bold": "Roboto-Bold",
        "heading": "Roboto-Bold",
        "data": "Roboto-Bold",
    }

    # Helvetica fallback when the Roboto/Bebas faces failed to register.
    _STYLE_TO_FALLBACK = {
        "value": "Helvetica",
        "body": "Helvetica",
        "category": "Helvetica",
        "patient": "Helvetica-Bold",
        "condensed": "Helvetica",
        "condensed_bold": "Helvetica-Bold",
    }

    def font(self, style="body"):
        """Return the registered font name for a semantic typography role."""
        if self.has_custom_fonts:
            return self._STYLE_TO_FACE.get(style, "Roboto-Light")
        return self._STYLE_TO_FALLBACK.get(style, "Helvetica-Bold")

    def _draw_clara_footer(self, c, clara_id, left=True, right=True):
        """Draw the repeating Clara ID footer — Roboto Black 12pt #080606."""
        if not clara_id:
            return
        c.setFillColor(NEAR_BLACK)
        c.setFont(self.font("footer"), 12)
        if left:
            c.drawString(*CLARA_ID_LEFT_POS, clara_id)
        if right:
            c.drawString(*CLARA_ID_RIGHT_POS, clara_id)

    def _draw_pill_label(self, c, label, pill_x, pill_cy, pill_w,
                         max_size=18, min_size=9, color=TEXT_WHITE):
        """Center a status-pill label in its capsule — Roboto Bold (#FFFFFF),
        18pt per the reference, auto-shrunk to fit the pill width."""
        font_name = self.font("pill")
        fs = max_size
        # Shrink until the label fits inside ~90% of the pill width.
        while fs > min_size and c.stringWidth(label, font_name, fs) > pill_w * 0.9:
            fs -= 0.5
        c.setFillColor(color)
        c.setFont(font_name, fs)
        tw = c.stringWidth(label, font_name, fs)
        c.drawString(pill_x + (pill_w - tw) / 2, pill_cy - fs * 0.32, label)

    def _draw_value_with_unit(self, c, value, unit, x, y, color,
                              value_size=38, unit_size=18, gap=3):
        """Draw a big numeric value + trailing unit on one baseline.
        Both are Roboto Light (clara_font_reference 'big numeric values' = 38pt,
        unit = 18pt); hemoglobin overrides value_size to 46pt."""
        value_font = self.font("value")
        c.setFillColor(color)
        c.setFont(value_font, value_size)
        c.drawString(x, y, value)
        if unit:
            vw = c.stringWidth(value, value_font, value_size)
            c.setFont(value_font, unit_size)
            c.drawString(x + vw + gap, y, unit)

    def _page_svg(self, page_num):
        # Girl reports use the `PageN_Girl.svg` background; boy/default uses `PageN.svg`.
        suffix = "_Girl" if self.is_girl else ""
        return os.path.join(self.backgrounds_root, f"Page {page_num}", f"Page{page_num}{suffix}.svg")

    def generate(self, data, output_path):
        c = pdf_canvas.Canvas(output_path, pagesize=PAGE_SIZE)

        # Page 1 — cover + student info (with data overlay)
        draw_background(c, self._page_svg(1))
        self._draw_page1_overlay(c, data)
        c.showPage()

        # Page 2 — BMI (left page data overlay; right "Summary" page background only)
        draw_background(c, self._page_svg(2))
        self._draw_page2_overlay(c, data)
        c.showPage()

        # Page 3 — Nervous System (left, no data yet) + Vitals (right) overlay
        draw_background(c, self._page_svg(3))
        self._draw_page3_overlay(c, data)
        c.showPage()

        # Page 4 — Anemia Screening (left) + Cardiovascular System (right)
        draw_background(c, self._page_svg(4))
        self._draw_page4_overlay(c, data)
        c.showPage()

        # Page 5 — Respiratory System (left) + Personal Hygiene (right)
        draw_background(c, self._page_svg(5))
        self._draw_page5_overlay(c, data)
        c.showPage()

        # Pages 6-8. Page 6 is background-only. Page 7 overlays ENT. Page 8 overlays
        # General Examination.
        for page_num in range(6, 9):
            svg = self._page_svg(page_num)
            if os.path.exists(svg):
                draw_background(c, svg)
                if page_num == 7:
                    self._draw_page7_overlay(c, data)
                elif page_num == 8:
                    self._draw_page8_overlay(c, data)
            else:
                c.setFont(self.font("regular"), 14)
                c.drawString(72, PAGE_HEIGHT - 100, f"Page {page_num} background missing")
            c.showPage()

        c.save()

    # ---------- Page 1: Cover ----------

    # Page-1 cover field anchors, as supplied by the layout editor in SVG viewBox
    # units (0..5700 x, 0..3900 y) with a BOTTOM-UP origin — so they convert to PDF
    # points by a plain * SCALE_X / * SCALE_Y (no y-flip). Paste new editor coords
    # straight in here; _svg_pt() does the conversion.
    PAGE1_ANCHORS = {
        "report_year":       (2912.3, 1833.4),  # health screening report year
        "date_of_screening": (2715.2, 1778.2),
        "school":            (2514.2, 1723.0),
        "clara_id":          (2526.1, 1663.9),
        "name":              (2498.5, 1529.9),
        "dob":               (2470.9, 1474.8),
        "sex":               (2459.1, 1419.6),
        "class":             (2553.7, 1321.1),  # "standard"
        "section":           (2533.9, 1264.3),  # "division"
        "roll_no":           (2514.2, 1206.0),
        "school_logo":       (2671.9, 2224.3),  # logo box CENTER
        "qr":                (2455.1, 975.1),   # qr box CENTER
    }

    @staticmethod
    def _svg_pt(x_svg, y_svg):
        """Convert an editor SVG-viewBox anchor (bottom-up) to PDF points."""
        return x_svg * SCALE_X, y_svg * SCALE_Y

    def _fetch_image(self, url):
        """Download a remote image (e.g. the school logo) into a PIL image.
        Returns None on any failure so the cover still renders without it."""
        if not url:
            return None
        try:
            resp = requests.get(url, timeout=10)
            resp.raise_for_status()
            return Image.open(BytesIO(resp.content))
        except Exception as e:
            print(f"⚠️  Could not load school logo from {url}: {e}")
            return None

    def _make_qr_image(self, content):
        """Render a QR code PIL image for the given content, or None if qrcode
        isn't installed. (Placeholder target until the real QR link is finalized.)"""
        if not _HAS_QR:
            return None
        qr = qrcode.QRCode(border=1, box_size=10)
        qr.add_data(content or "CLARA")
        qr.make(fit=True)
        return qr.make_image(fill_color="black", back_color="white").convert("RGB")

    def _draw_image_centered(self, c, pil_img, cx, cy, box_w, box_h):
        """Draw a PIL image fit inside box_w x box_h, centered on (cx, cy)."""
        iw, ih = pil_img.size
        if not iw or not ih:
            return
        scale = min(box_w / iw, box_h / ih)
        w, h = iw * scale, ih * scale
        c.drawImage(ImageReader(pil_img), cx - w / 2, cy - h / 2, width=w, height=h,
                    preserveAspectRatio=True, mask="auto")

    def _draw_page1_overlay(self, c, data):
        """Overlay cover (Spread 1) data: screening metadata, student identity,
        the school logo (fetched from the backend URL), and a QR code.

        All field positions live in PAGE1_ANCHORS as editor SVG-viewBox coords;
        _svg_pt() converts each to PDF points.
        """
        student = data.get("student", {})
        clara_id = student.get("clara_id", "")
        school_name = data.get("camp_name", "")
        logo_url = data.get("school_logo_url", "")

        today = datetime.now().strftime("%d/%m/%Y")
        screening_date = data.get("screening_date") or today
        report_year = data.get("report_year") or str(datetime.now().year)

        def P(key):
            return self._svg_pt(*self.PAGE1_ANCHORS[key])

        # ── Screening metadata + school + clara id — Roboto Bold 12pt #B2E2F2 ──
        c.setFillColor(HEADING_BLUE)
        c.setFont(self.font("label"), 12)
        c.drawString(*P("report_year"), str(report_year))
        c.drawString(*P("date_of_screening"), screening_date)
        if school_name:
            c.drawString(*P("school"), str(school_name))
        if clara_id:
            c.drawString(*P("clara_id"), str(clara_id))

        # ── Student identity block — Bebas Neue 16pt #8EC8D1 ──
        c.setFillColor(PATIENT_TEAL)
        c.setFont(self.font("patient"), 16)
        for key in ("name", "dob", "sex", "class", "section", "roll_no"):
            value = student.get(key, "")
            if value:
                c.drawString(*P(key), str(value))

        # ── School logo — fetched from the backend URL, centered on its anchor ──
        logo_img = self._fetch_image(logo_url)
        if logo_img is not None:
            self._draw_image_centered(c, logo_img, *P("school_logo"), 90, 90)

        # ── QR code — placeholder content (clara id) until the real link is set ──
        qr_img = self._make_qr_image(clara_id or "CLARA")
        if qr_img is not None:
            self._draw_image_centered(c, qr_img, *P("qr"), 80, 80)

    # ---------- Page 2: BMI ----------

    def _draw_wrapped(self, c, text, x, y, max_width, font_name, font_size, leading, color=colors.white, max_lines=None):
        """Draw left-aligned wrapped text downward from baseline y. Returns the y after the last line.

        Honors explicit newlines in `text`; wraps each paragraph to max_width.
        Empty paragraphs (e.g. from "\\n\\n") render as a blank line gap.
        """
        c.setFont(font_name, font_size)
        c.setFillColor(color)
        cur_y = y
        lines_drawn = 0
        for paragraph in text.split("\n"):
            if not paragraph.strip():
                # blank line — advance by one leading without drawing
                cur_y -= leading
                lines_drawn += 1
                if max_lines and lines_drawn >= max_lines:
                    return cur_y
                continue
            words = paragraph.split()
            line = ""
            for word in words:
                trial = (line + " " + word).strip()
                if c.stringWidth(trial, font_name, font_size) <= max_width:
                    line = trial
                else:
                    if line:
                        c.drawString(x, cur_y, line)
                        cur_y -= leading
                        lines_drawn += 1
                        if max_lines and lines_drawn >= max_lines:
                            return cur_y
                    line = word
            if line:
                c.drawString(x, cur_y, line)
                cur_y -= leading
                lines_drawn += 1
                if max_lines and lines_drawn >= max_lines:
                    return cur_y
        return cur_y

    def _draw_page2_overlay(self, c, data):
        """Overlay BMI data on the left page of Spread 2.

        Left page background sits at SVG (x=369, y=196) — PDF x≈88..684.
        Coordinates below are first-pass estimates (read off a 20-pt grid on
        the rendered page) and will be refined with editor coords.

        Renders: Height/Weight/BMI values, the Excel conclusion (col J) and
        Fact (col K) for the WHO BMI category. Observation left blank; category
        highlight pill + weight pill + scale marker are deferred (coords pending).
        """
        measurements = data.get("measurements", {})
        category = data.get("bmi_category", "")
        clara_id = data.get("student", {}).get("clara_id", "")
        ref = self.bmi_reference().get(category.lower(), {})

        # ── Height / Weight / BMI values (after their labels, top-left) ──
        # Roboto Bold 12pt #B2E2F2 (clara_font_reference, Page 2:
        # "Height: XXX cm / Weight: XX kg / BMI:" = Roboto Bold 12pt #B2E2F2).
        c.setFillColor(HEADING_BLUE)
        c.setFont(self.font("label"), 12)
        height = measurements.get("height", "")
        weight = measurements.get("weight", "")
        bmi = measurements.get("bmi", "")
        if height:
            c.drawString(200, 690, f"{height} cm")
        if weight:
            c.drawString(200, 665, f"{weight} kg")
        if bmi:
            c.drawString(179.1, 640.9, str(bmi))

        # ── Weight value + "kg" unit inside the bottom "Weight: __ kg" capsule ──
        # Same Roboto Bold 12pt #B2E2F2 as the weight value above (only "Weight:"
        # is baked into the capsule art; we overlay "<weight> kg").
        if weight:
            c.drawString(522.1, 318.1, f"{weight} kg")

        # ── Conclusion (Excel col J), confined to the editor-marked box:
        #    x 146.3..639.4 (w≈493), y 291.6 (top) .. 238.6 (bottom), h≈53.
        conclusion = ref.get("conclusion", "")
        if conclusion:
            self._draw_wrapped(
                c, conclusion, x=146, y=284, max_width=493,
                font_name=self.font("body"), font_size=8.8, leading=10.5,
                color=TEXT_WHITE, max_lines=5,
            )

        # ── Fact (Excel col K), confined to the editor-marked box:
        #    x 119.8..370.8 (w≈251), y 200.3 (top) .. 157.2 (bottom), h≈43.
        #    The black background is part of the page art — do NOT mask here.
        fact = ref.get("fact", "")
        if fact:
            fact_body, _, fact_source = fact.partition("Source")
            # Fact callout — Roboto Black #FFFFFF (reference nominal 12pt; sized
            # to 9pt to fit the editor-marked box, x122..368 / h≈43).
            self._draw_wrapped(
                c, "Fact: " + fact_body.strip(), x=122, y=193, max_width=246,
                font_name=self.font("fact"), font_size=9, leading=10.5,
                color=TEXT_WHITE, max_lines=3,
            )
            if fact_source.strip():
                # Source attribution — Roboto Light #FFFFFF.
                self._draw_wrapped(
                    c, ("Source" + fact_source).strip(), x=122, y=164, max_width=246,
                    font_name=self.font("body"), font_size=7, leading=8.5,
                    color=TEXT_WHITE, max_lines=1,
                )

        # ── BMI Observation capsule (label baked into the art) ──
        # Only the Normal capsule exists, so it shows "Normal".
        self._draw_pill_capsule(c, OBSERVATION_NORMAL_SVG, 254.7, 373.6)

        # ── Diet strip (left page) keyed by BMI category ──
        diet_svg = BMI_DIET_SVG.get(category)
        if diet_svg:
            self._draw_diet(c, os.path.join("Page 2", diet_svg), LEFT_DIET_ANCHOR)

        # ── Bottom Clara IDs (left + right pages of the spread) ──
        self._draw_clara_footer(c, clara_id)

        # ── Right page: Summary bar chart markers ──
        self._draw_page2_summary(c, data)

    def _draw_pill(self, c, x, y, w, h, text, fill_color=colors.HexColor("#2D7B43"),
                   text_color=colors.white, font_size=12, font_style="data"):
        """Draw a rounded-end 'capsule' with centered text. (x, y) = bottom-left."""
        c.setFillColor(fill_color)
        c.setStrokeColor(fill_color)
        c.roundRect(x, y, w, h, h / 2, fill=1, stroke=0)
        font_name = self.font(font_style)
        c.setFillColor(text_color)
        c.setFont(font_name, font_size)
        tw = c.stringWidth(text, font_name, font_size)
        baseline_y = y + h / 2 - font_size * 0.3
        c.drawString(x + (w - tw) / 2, baseline_y, text)

    def _draw_summary_line(self, c, x_end, y):
        """Draw a white horizontal progress line from the bar's left edge to x_end."""
        c.setStrokeColor(colors.white)
        c.setLineWidth(SUMMARY_LINE_THICKNESS)
        c.setLineCap(1)  # round caps look cleaner on the gradient bar
        c.line(SUMMARY_BAR_LEFT_X, y, x_end, y)

    def _draw_page2_summary(self, c, data):
        """Place rating lines on the right-page Summary bar chart.

        For each rated system row, draws a white horizontal line from the bar's
        left edge to the rating column's x position. Rows without data are skipped.
        """
        ratings = {
            "bmi": bmi_summary_rating(data.get("bmi_category", "")),
            "hemoglobin": hemoglobin_summary_rating(data.get("blood_work", {})),
            "nervous": nervous_summary_rating(data.get("nervous", {})),
            "cardio": cardio_summary_rating(data.get("cardio", {})),
            "respiratory": respiratory_summary_rating(data.get("respiratory", {})),
            "ent": ent_summary_rating(data.get("ent", {})),
            "dental": dental_summary_rating(
                {k: (v or {}).get("value") for k, v in (data.get("dental", {}) or {}).items()}
            ),
            # Spirometry / Eye Vision have no inputs nor defaults in
            # the current JSON contract; left None.
        }
        for row_key, rating in ratings.items():
            if not rating:
                continue
            x_end = SUMMARY_X.get(rating)
            y = SUMMARY_ROW_Y.get(row_key)
            if x_end is not None and y is not None:
                self._draw_summary_line(c, x_end, y)

    # ---------- Page 3: Nervous System (left) + Vitals (right) ----------

    def _draw_page3_overlay(self, c, data):
        """Page 3 spread.

        Left ("Nervous System"): no JSON data yet for mental/motor/reflexes,
        so only the Clara ID strip is populated.

        Right ("VITALS"): pulse + oxymetry values in their boxes, short
        observation label, and the matching outlook+actions paragraph from
        the VITALS Excel sheet J4 (picked by actual readings).

        Coordinates are first-pass estimates and will be refined with editor coords.
        """
        student = data.get("student", {})
        clara_id = student.get("clara_id", "")
        vitals = data.get("vitals", {})
        pulse = vitals.get("pulse_rate", "")
        oxygen = vitals.get("oxymetry", "")

        # ── Right page (VITALS) ──
        c.setFillColor(colors.white)

        # Pulse Rate big numeric value — Roboto Light 38pt #B2E2F2 ("bpm" unit
        # is baked into the box art). clara_font_reference, Page 3 VITALS.
        if pulse:
            self._draw_value_with_unit(c, str(pulse), "", 780.6, 541.3, HEADING_BLUE)

        # Oxymetry big numeric value + "%" unit — Roboto Light 38pt / 18pt #B2E2F2.
        if oxygen:
            self._draw_value_with_unit(c, str(oxygen), "%", 783.1, 440.4, HEADING_BLUE)

        # Conclusion (Outlook + Actions) confined to the editor-marked box:
        # x 731.4..1017.7 (w≈286), y 239.9 (bottom) .. 291.6 (top), h≈52.
        vitals_ref = self.vitals_reference()
        conclusion = vitals_pick_section(vitals_ref, pulse, oxygen) if vitals_ref else ""
        if conclusion:
            conclusion = conclusion.replace("\nActions:", "\n\nActions:")
            self._draw_wrapped(
                c, conclusion, x=731.4, y=285, max_width=286,
                font_name=self.font("body"), font_size=8.8, leading=10,
                color=TEXT_WHITE, max_lines=5,
            )

        # VITALS Observation capsule (label baked into the art). Only the Normal
        # capsule exists, so it shows "Normal".
        self._draw_pill_capsule(c, OBSERVATION_NORMAL_SVG, 842.4, 367.3)

        # ── Diet strip (right page, Vitals) ──
        self._draw_diet(c, os.path.join("Page 3", VITALS_DIET_SVG), RIGHT_DIET_ANCHOR)

        # Right-page Clara ID bottom strip
        self._draw_clara_footer(c, clara_id, left=False)

        # ── Left page (Nervous System) ──
        self._draw_page3_nervous(c, data)

        self._draw_clara_footer(c, clara_id, right=False)

    def _draw_page3_nervous(self, c, data):
        """Left page: overlay Mental Status / Motor Strength / Reflexes pill + comment.

        Each section uses the SVG pill from backgrounds/Boy/Page 3/<sub>/<Cat>.svg
        and the corresponding comment from the NERVOUS SYSTEM Excel sheet (column H).
        Until the JSON gains those fields, defaults from NERVOUS_DEFAULTS are used.
        """
        nervous_data = data.get("nervous", {}) or {}
        ref = self.nervous_reference()

        # (sub_param_key, pill_left_x, pill_center_y, pill_w, pill_h, comment_max_lines)
        # The editor's FIELDS panel gives (x, y) where x = pill left edge and
        # y = pill VERTICAL CENTER (so the pill aligns visually with the section
        # heading's baseline gap).
        rows = [
            ("mental_status",  156.4, 598.0, 170, 46, 6),
            ("motor_strength", 160.2, 463.1, 170, 46, 6),
            ("reflexes",       160.2, 316.8, 170, 46, 6),
        ]
        for key, pill_x, pill_cy, w, h, com_lines in rows:
            category = nervous_data.get(key) or NERVOUS_DEFAULTS.get(key, "")
            if not category:
                continue
            pill_y = pill_cy - h / 2
            svg_name = NERVOUS_PILL_SVGS.get(key, {}).get(category)
            if svg_name:
                self._draw_overlay_svg(c, os.path.join("Page 3", svg_name), pill_x, pill_y, w, h)
                label = os.path.splitext(os.path.basename(svg_name))[0]
                self._draw_pill_label(c, label, pill_x, pill_cy, w)
            comment = (ref.get(key, {}) or {}).get(category, "")
            if comment:
                self._draw_wrapped(
                    c, comment, x=pill_x + 4, y=pill_y - 14, max_width=170,
                    font_name=self.font("body"), font_size=8.8, leading=12,
                    color=TEXT_WHITE, max_lines=com_lines,
                )

    # ---------- Page 4: Anemia Screening (left) + Cardiovascular System (right) ----------

    def _draw_page4_overlay(self, c, data):
        """Page 4 spread.

        LEFT (Anemia Screening): hemoglobin value, category pill, comment from
        ANEMIA Excel J column. Category derived from hemoglobin text since the
        API returns descriptive text (no numeric g/dl value).

        RIGHT (Cardiovascular): Heart Sounds / Capillary Refill / Pulse Quality
        capsule SVG overlays + comment from CARDIOVASCULAR SYSTEM sheet column H.
        Defaults to the "positive" category for each sub-param until JSON gains
        cardiovascular fields under data.cardio.
        """
        student = data.get("student", {})
        clara_id = student.get("clara_id", "")
        blood_work = data.get("blood_work", {})
        cardio_data = data.get("cardio", {}) or {}

        # ============ LEFT: Anemia Screening ============
        anemia_cat = anemia_category_from_blood_work(blood_work)
        anemia_ref = self.anemia_reference()
        anemia_row = anemia_ref.get(anemia_cat.lower(), {})

        # Hemoglobin value box — API returns text not numeric, so we use the
        # blood_work numeric field if present, otherwise leave blank.
        hb_numeric = blood_work.get("hemoglobin_numeric") or ""
        if hb_numeric:
            # Large hemoglobin value — Roboto Light 46pt #8EC8D1 ("g/dl" unit is
            # baked into the box art). clara_font_reference, Page 4 Anemia.
            self._draw_value_with_unit(
                c, str(hb_numeric), "", 187.9, 610.6 - 6, PATIENT_TEAL, value_size=46
            )

        # Comment text (Excel J - Parental Guidance) below the threshold table.
        comment = anemia_row.get("comment", "")
        if comment:
            self._draw_wrapped(
                c, comment, x=151.3, y=314.3, max_width=437,
                font_name=self.font("body"), font_size=8.8, leading=11,
                color=TEXT_WHITE, max_lines=8,
            )

        # Severity arc + marker dot on the blood-cell circle.
        self._draw_anemia_severity_arc(c, anemia_cat)

        # ============ RIGHT: Cardiovascular System ============
        cardio_ref = self.cardio_reference()
        # (sub_key, pill_left_x, pill_center_y, pill_w, pill_h, comment_lines)
        cardio_rows = [
            ("heart_sounds",      741.5, 616.9, 170, 46, 4),
            ("capillary_refill",  745.3, 453.0, 170, 46, 4),
            ("pulse_quality",     745.3, 287.8, 170, 46, 4),
        ]
        for key, pill_x, pill_cy, w, h, com_lines in cardio_rows:
            category = cardio_data.get(key) or CARDIO_DEFAULTS.get(key, "")
            if not category:
                continue
            pill_y = pill_cy - h / 2
            svg_name = CARDIO_PILL_SVGS.get(key, {}).get(category)
            if svg_name:
                self._draw_overlay_svg(c, os.path.join("Page 4", svg_name), pill_x, pill_y, w, h)
                # Short label = SVG filename minus extension
                label = os.path.splitext(os.path.basename(svg_name))[0]
                self._draw_pill_label(c, label, pill_x, pill_cy, w)
            comment = (cardio_ref.get(key, {}) or {}).get(category, "")
            if comment:
                self._draw_wrapped(
                    c, comment, x=pill_x + 4, y=pill_y - 14, max_width=190,
                    font_name=self.font("body"), font_size=8.8, leading=11,
                    color=TEXT_WHITE, max_lines=com_lines,
                )

        # Anemia Observation capsule (LEFT) + Cardiovascular Doctor Visit (RIGHT).
        # No cardio doctor-visit field in the JSON yet -> defaults to NO.
        self._draw_pill_capsule(c, OBSERVATION_NORMAL_SVG, 258.5, 556.4)
        self._draw_pill_capsule(c, DOCTOR_VISIT_SVG["NO"], 1010.1, 273.9)

        # ── Diet strip (left page, Anemia) keyed by anemia category ──
        diet_svg = ANEMIA_DIET_SVG.get(anemia_cat.lower())
        if diet_svg:
            self._draw_diet(c, os.path.join("Page 4", diet_svg), LEFT_DIET_ANCHOR)

        # Clara ID strips at the bottom of both pages
        self._draw_clara_footer(c, clara_id)

    # Circle geometry for the Page-4 blood-cell severity dial (left page).
    # Categories sit at the 4 cardinal positions on a dashed circle.
    _ANEMIA_CIRCLE = {"cx": 535.0, "cy": 432.0, "r": 72.0}
    _ANEMIA_ANGLES = {
        # math degrees: 0=right, 90=top, 180=left, 270=bottom
        "severe anemic": 0.0,            # right
        "moderate anemic": 90.0,         # top
        "mild anemic": 180.0,            # left
        "normal ( non anemic)": 270.0,   # bottom
    }

    # Arc sweep per category. Arc starts at Normal (bottom, 270°) and grows
    # clockwise toward the patient's category — so a longer arc means a worse
    # reading. Normal = 0° (just the marker dot, no arc).
    _ANEMIA_ARC_SWEEP = {
        "normal ( non anemic)": 0,      # dot only at Normal
        "mild anemic":         -90,     # CW from Normal -> Mild (left)
        "moderate anemic":     -180,    # CW from Normal -> Moderate (top)
        "severe anemic":       -270,    # CW from Normal -> Severe (right)
    }

    def _draw_anemia_severity_arc(self, c, category):
        """Draw the dynamic severity arc + marker dot on the blood-cell circle.

        Arc starts at the Normal position (bottom of the circle) and sweeps
        clockwise to the patient's category — so a healthy reading shows just
        a dot at Normal, and the arc lengthens as severity increases.
        """
        cat_lower = category.lower()
        end_angle = self._ANEMIA_ANGLES.get(cat_lower)
        sweep = self._ANEMIA_ARC_SWEEP.get(cat_lower)
        if end_angle is None or sweep is None:
            return
        cx = self._ANEMIA_CIRCLE["cx"]
        cy = self._ANEMIA_CIRCLE["cy"]
        r = self._ANEMIA_CIRCLE["r"]

        # Arc from Normal (270°) CW to the patient's category position
        if sweep != 0:
            c.saveState()
            c.setStrokeColor(colors.HexColor("#B5E853"))  # lime-green progress arc
            c.setLineWidth(3)
            c.setLineCap(1)
            path = c.beginPath()
            start_x = cx + r * math.cos(math.radians(270))
            start_y = cy + r * math.sin(math.radians(270))
            path.moveTo(start_x, start_y)
            path.arcTo(cx - r, cy - r, cx + r, cy + r, startAng=270, extent=sweep)
            c.drawPath(path, stroke=1, fill=0)
            c.restoreState()

        # Marker dot at the patient's category position
        rad = math.radians(end_angle)
        dot_x = cx + r * math.cos(rad)
        dot_y = cy + r * math.sin(rad)
        c.saveState()
        c.setFillColor(colors.HexColor("#2D7B43"))  # forest green (matches capsules)
        c.setStrokeColor(colors.white)
        c.setLineWidth(1.5)
        c.circle(dot_x, dot_y, 4, stroke=1, fill=1)
        c.restoreState()

    # ---------- Page 5: Respiratory System (left) + Personal Hygiene (right) ----------

    def _draw_page5_overlay(self, c, data):
        """Page 5 spread.

        LEFT (Respiratory): Breath Sounds / Effort / Cough capsules + comments.
        Doctor Visit pill at bottom-left. JSON has no respiratory data yet, so
        positive defaults are used.

        RIGHT (Personal Hygiene): Hair + Nail capsules with the category label,
        and a marker dot on the Poor/Moderate/Excellent scale below each.
        """
        student = data.get("student", {})
        clara_id = student.get("clara_id", "")
        respiratory_data = data.get("respiratory", {}) or {}
        hygiene_data = data.get("hygiene", {}) or {}

        # ============ LEFT: Respiratory System ============
        resp_ref = self.respiratory_reference()
        # (sub_key, pill_left_x, pill_center_y, pill_w, pill_h, com_max_lines, com_max_width)
        # Editor-marked comment zones (x left..right, y bottom..top):
        #   breath_sounds : x 141.2..283.7, y 527.4..580.4
        #   effort        : x 459.0..619.2, y 539.5..581.1
        #   cough         : x 141.2..283.7, y 374.3..423.5
        resp_rows = [
            ("breath_sounds", 141.2, 601.8, 170, 46, 4, 138),
            ("effort",        459.0, 604.3, 170, 46, 4, 156),
            ("cough",         141.2, 448.0, 170, 46, 4, 138),
        ]
        for key, pill_x, pill_cy, w, h, com_lines, com_max_w in resp_rows:
            category = respiratory_data.get(key) or RESPIRATORY_DEFAULTS.get(key, "")
            if not category:
                continue
            pill_y = pill_cy - h / 2
            svg_name = RESPIRATORY_PILL_SVGS.get(key, {}).get(category)
            if svg_name:
                self._draw_overlay_svg(c, os.path.join("Page 5", svg_name), pill_x, pill_y, w, h)
                label = os.path.splitext(os.path.basename(svg_name))[0]
                self._draw_pill_label(c, label, pill_x, pill_cy, w)
            comment = (resp_ref.get(key, {}) or {}).get(category, "")
            if comment:
                self._draw_wrapped(
                    c, comment, x=pill_x + 4, y=pill_y - 14, max_width=com_max_w,
                    font_name=self.font("body"), font_size=8.8, leading=11,
                    color=TEXT_WHITE, max_lines=com_lines,
                )

        # Doctor Visit Recommended capsule — no respiratory doctor-visit field in
        # the JSON yet, so it defaults to NO. Label is baked into the page art.
        self._draw_pill_capsule(c, DOCTOR_VISIT_SVG["NO"], 129.9, 253.8)

        # ── Diet strip (left page, Respiratory) ──
        self._draw_diet(c, os.path.join("Page 5", RESPIRATORY_DIET_SVG), LEFT_DIET_ANCHOR)

        # ============ RIGHT: Personal Hygiene ============
        hygiene_ref = self.hygiene_reference()
        hair_category = hygiene_category_from_text(
            hygiene_data.get("hair_value"), hygiene_data.get("hair_comment")
        )
        nail_category = hygiene_category_from_text(
            hygiene_data.get("nail_value"), hygiene_data.get("nail_comment")
        )

        # Hair + Nail Hygiene observation capsules (category text — Excellent /
        # Moderate / Needs Attention — is baked into each capsule SVG).
        hair_cap = os.path.join("Page 5", "Hair Hygiene", f"{hair_category}.svg")
        self._draw_pill_capsule(c, hair_cap, 752.9, 609.4)

        # Nail Hygiene folder is currently empty -> fall back to the Hair Hygiene
        # capsule of the same category (visually identical) until nail SVGs exist.
        nail_cap = os.path.join("Page 5", "Nail Hygiene", f"{nail_category}.svg")
        if not os.path.exists(os.path.join(self.backgrounds_root, nail_cap)):
            nail_cap = os.path.join("Page 5", "Hair Hygiene", f"{nail_category}.svg")
        self._draw_pill_capsule(c, nail_cap, 1051.7, 391.2)

        # Hair Hygiene scale marker
        # 3-point scale dot positions for hair: poor=750.3, moderate=877.7, excellent=1026.5, all y=516.1
        hair_scale = {
            "Needs Attention": (750.3, 532),
            "Moderate":        (877.7, 532),
            "Excellent":       (1026.5, 532),
        }
        self._draw_scale_marker(c, *hair_scale[hair_category])

        # Nail Hygiene scale marker
        # 3-point scale dot positions for nail: poor=926.9, moderate=1054.2, excellent=1203.1, all y=284.0
        nail_scale = {
            "Needs Attention": (926.9, 294),
            "Moderate":        (1054.2, 294),
            "Excellent":       (1203.1, 294),
        }
        self._draw_scale_marker(c, *nail_scale[nail_category])

        # Combined parental guidance — now confined to the LEFT half of the bottom
        # band so the diet strip can sit on the right. The Excel sheet has ONE shared
        # guidance per category for hair + nail, so we show the WORSE of the two.
        # Editor-marked comment box: x 747.8..968.5 (w≈221), y 116.3 (bottom)..199.5 (top).
        severity = {"Excellent": 0, "Moderate": 1, "Needs Attention": 2}
        worst = max((hair_category, nail_category), key=lambda k: severity.get(k, 0))
        combined_comment = hygiene_ref.get(worst, "")
        if combined_comment:
            self._draw_wrapped(
                c, combined_comment, x=747.8, y=193, max_width=220,
                font_name=self.font("body"), font_size=8.8, leading=11,
                color=TEXT_WHITE, max_lines=7,
            )

        # ── Diet strip (right page, Personal Hygiene) keyed by the worse category ──
        diet_svg = HYGIENE_DIET_SVG.get(worst)
        if diet_svg:
            self._draw_diet(c, os.path.join("Page 5", diet_svg), RIGHT_DIET_ANCHOR)

        # Clara ID strips
        self._draw_clara_footer(c, clara_id)

    def _scale_triangle_drawing(self):
        """Lazy-load and cache the white-recolored triangle SVG marker."""
        if getattr(self, "_triangle_dwg", None) is not None:
            return self._triangle_dwg
        # Triangle SVG sits at the repo's backgrounds/ root.
        svg_path = os.path.join(
            os.path.dirname(self.backgrounds_root),
            "triangle-shape-container-empty-wrapper-point-svgrepo-com.svg",
        )
        if not os.path.exists(svg_path):
            self._triangle_dwg = None
            return None
        d = svg2rlg(svg_path)
        # Recolor outline/fill to white — but only touch elements that already
        # have a non-transparent color (skip background `rect`s with fill=none).
        white = colors.white
        def walk(node):
            if getattr(node, "strokeColor", None) is not None:
                try:
                    node.strokeColor = white
                except Exception:
                    pass
            if getattr(node, "fillColor", None) is not None:
                try:
                    node.fillColor = white
                except Exception:
                    pass
            children = getattr(node, "contents", None)
            if children:
                for child in children:
                    walk(child)
        walk(d)
        self._triangle_dwg = d
        return d

    def _draw_scale_marker(self, c, x, y, size=14):
        """Draw the triangle marker SVG centered at (x, y) on a hygiene scale track."""
        d = self._scale_triangle_drawing()
        if d is None:
            # Fallback: small filled circle if SVG missing
            c.saveState()
            c.setFillColor(colors.white)
            c.circle(x, y, 4, fill=1, stroke=0)
            c.restoreState()
            return
        # The svglib Drawing carries a persistent scale, so build a fresh copy each call.
        from copy import deepcopy
        dwg = deepcopy(d)
        scale = size / dwg.width
        dwg.scale(scale, scale)
        dwg.width *= scale
        dwg.height *= scale
        renderPDF.draw(dwg, c, x - size / 2, y - size / 2)

    # ---------- Page 7: Dental Examination (left) + ENT Examination (right) ----------

    def _draw_page7_overlay(self, c, data):
        """Page 7 spread — Dental Examination on the LEFT page, ENT on the RIGHT.

        ENT has 4 sub-parameters arranged in a 2x2 grid on the right page:
        ears + nose on the top row, hearing_status + throat on the bottom.
        A single comment box below combines all four observations. JSON has no
        ENT fields yet, so positive defaults are used.

        Editor-marked pill positions (x = left edge, y = vertical center):
          ears           : (739.0,  693.9)
          nose           : (1085.8, 693.9)
          hearing_status : (739.0,  354.6)
          throat         : (1099.6, 354.6)
          comment (top-left of combined box): (776.8, 256.3)
        """
        student = data.get("student", {})
        clara_id = student.get("clara_id", "")
        ent_data = data.get("ent", {}) or {}
        ent_ref = self.ent_reference()

        # (sub_key, pill_x, pill_cy, pill_w, pill_h, com_x, com_y, com_max_w, com_lines)
        # Editor-marked comment zones (x left..right, y bottom..top):
        #   ears           : x 742.8..882.7,    y 618.5..663.6
        #   nose           : x 1089.6..1232.1,  y 618.2..663.6
        #   hearing_status : x 742.8..882.7,    y 295.4..324.4  (only ~3 lines fit)
        #   throat         : x 1104.7..1243.4,  y 284.3..324.6
        ent_rows = [
            ("ears",            739.0, 693.9, 170, 46,  742.8, 657, 139, 4),
            ("nose",           1085.8, 693.9, 170, 46, 1089.6, 657, 142, 4),
            ("hearing_status",  739.0, 354.6, 170, 46,  742.8, 318, 139, 3),
            ("throat",         1099.6, 354.6, 170, 46, 1104.7, 318, 138, 4),
        ]
        for key, pill_x, pill_cy, w, h, com_x, com_y, com_max_w, com_lines in ent_rows:
            category = ent_data.get(key) or ENT_DEFAULTS.get(key, "")
            if not category:
                continue
            pill_y = pill_cy - h / 2
            svg_name = ENT_PILL_SVGS.get(key, {}).get(category)
            if svg_name:
                self._draw_overlay_svg(c, os.path.join("Page 7", svg_name), pill_x, pill_y, w, h)
                label = os.path.splitext(os.path.basename(svg_name))[0]
                self._draw_pill_label(c, label, pill_x, pill_cy, w)
            entry = (ent_ref.get(key, {}) or {}).get(category, {}) or {}
            comment = entry.get("comment", "")
            if comment:
                self._draw_wrapped(
                    c, comment, x=com_x, y=com_y, max_width=com_max_w,
                    font_name=self.font("body"), font_size=8.8, leading=11,
                    color=TEXT_WHITE, max_lines=com_lines,
                )

        # Bottom "Comment" dark box is baked into the template SVG — no overlay needed.

        # ── Diet strip (right page, ENT) ──
        self._draw_diet(c, os.path.join("Page 7", ENT_DIET_SVG), RIGHT_DIET_ANCHOR)

        # Dental overlay on the LEFT page of the spread.
        self._draw_page7_dental(c, data)

        # Clara ID strips at the bottom of both pages (Dental left, ENT right)
        self._draw_clara_footer(c, clara_id)

    def _draw_page7_dental(self, c, data):
        """Page 7 LEFT page — Dental Examination overlay.

        Seven sub-parameter pills in a 2-column grid + Dentist Visit Recommended
        pill at bottom-left.

        Editor-marked pill positions (x = pill LEFT edge, y = pill VERTICAL CENTER):
          LEFT  col: dental_cavity          (240.9, 632.1)
                     gum_health             (240.9, 531.2)
                     alignment              (240.9, 426.5)
                     other_condition        (240.9, 334.5)
          RIGHT col: nursing_bottle_caries  (547.3, 633.3)
                     dental_fluorosis       (547.3, 530.4)
                     oral_hygiene           (547.3, 425.8)

        Each pill's comment text starts ~35pt below the pill center, left-aligned
        with the column's label start, matching the Page 8 spacing.
        """
        dental_data = data.get("dental", {}) or {}
        ref = self.dental_reference()

        # Reduced from 170x46 so the right-column pills (pill_x=547.3) don't
        # bleed past the page midline (~684) into the ENT page on the right.
        pill_w, pill_h = 141, 36
        # (key, pill_x, pill_cy, w, h, com_x, com_y, com_max_w, com_lines)
        rows = [
            # LEFT column — comment x aligned to each section heading's x.
            ("dental_cavity",         240.9, 632.1, pill_w, pill_h, 129.9, 597, 220, 3),
            ("gum_health",            240.9, 531.2, pill_w, pill_h, 129.9, 496, 220, 3),
            ("alignment",             240.9, 426.5, pill_w, pill_h, 129.9, 392, 220, 3),
            ("other_condition",       240.9, 334.5, pill_w, pill_h, 133.7, 300, 220, 3),
            # RIGHT column — comment x aligned to each section heading's x.
            ("nursing_bottle_caries", 547.3, 633.3, pill_w, pill_h, 417.4, 598, 220, 3),
            ("dental_fluorosis",      547.3, 530.4, pill_w, pill_h, 422.5, 495, 220, 3),
            ("oral_hygiene",          547.3, 425.8, pill_w, pill_h, 426.2, 391, 220, 3),
        ]
        for key, pill_x, pill_cy, w, h, com_x, com_y, com_max_w, com_lines in rows:
            raw = (dental_data.get(key) or {}).get("value")
            category = dental_normalize(key, raw) or DENTAL_DEFAULTS.get(key, "")
            if not category:
                continue
            pill_y = pill_cy - h / 2
            svg_name = DENTAL_PILL_SVGS.get(key, {}).get(category)
            if svg_name:
                self._draw_overlay_svg(c, os.path.join("Page 7", svg_name), pill_x, pill_y, w, h)
                label = DENTAL_PILL_LABELS.get((key, category)) or \
                    os.path.splitext(os.path.basename(svg_name))[0].replace("-", " ")
                self._draw_pill_label(c, label, pill_x, pill_cy, w)
            comment = (ref.get(key, {}) or {}).get(category, "")
            if comment:
                self._draw_wrapped(
                    c, comment, x=com_x, y=com_y, max_width=com_max_w,
                    font_name=self.font("body"), font_size=8.8, leading=11,
                    color=TEXT_WHITE, max_lines=com_lines,
                )

        # Dentist Visit Recommended capsule (YES/NO) — label baked into the art.
        dv = dental_visit_label_from_text((dental_data.get("dental_visit") or {}).get("value"))
        self._draw_pill_capsule(c, DOCTOR_VISIT_SVG[dv], 140.0, 188.2)

    # ---------- Page 8: General Examination ----------

    def _draw_page8_overlay(self, c, data):
        """Page 8 — General Examination spread (single content area, two columns).

        Editor-marked pill positions (x = pill left edge, y = pill vertical center).
        Each pill gets its category capsule SVG overlay + the matching observation
        comment (col H) drawn just below. A Doctor Visit Recommended pill sits at
        the bottom-right driven by the GENERAL EXAMINATION doctor_visit text.

        Posture and gait_and_coordination are not present in the current API
        contract, so they fall back to positive defaults from GENERAL_DEFAULTS.
        The center-bottom description paragraph is part of the template SVG.
        """
        student = data.get("student", {})
        clara_id = student.get("clara_id", "")
        general = data.get("general", {}) or {}
        ref = self.general_examination_reference()

        # (field_key, pill_x, pill_cy, pill_w, pill_h, com_x, com_y, com_max_w, com_lines)
        # Comment placement is a first-pass estimate (offset below each pill) —
        # refine with editor zone corners as you did for ENT.
        # Page-8 pill SVGs are 495x149 (visible pill ~88% w x 62% h). Render at 176x52
        # so the visible pill is ~155x32, matching the ENT/cardio pill SVGs (684x186).
        pill_w, pill_h = 176, 52
        rows = [
            # LEFT column — editor-marked comment zones (x 152.6..365.7, w≈213):
            #   pallor          : y 658.3..693.9
            #   icterus         : y 557.7..594.2
            #   cyanosis        : y 456.5..493.1
            #   clubbing        : y 362.0..393.5
            #   lympha_denopathy: y 258.8..294.1
            #   skin_assessment : y 164.2..194.5
            ("pallor",                276.2, 722.9, pill_w, pill_h, 145.6, 687, 213, 3),
            ("icterus",               276.2, 620.7, pill_w, pill_h, 145.6, 587, 213, 3),
            ("cyanosis",              276.2, 521.1, pill_w, pill_h, 145.6, 486, 213, 3),
            ("clubbing",              276.2, 421.5, pill_w, pill_h, 145.6, 386, 213, 3),
            ("lympha_denopathy",      276.2, 320.6, pill_w, pill_h, 145.6, 287, 213, 3),
            ("skin_assessment",       276.2, 222.2, pill_w, pill_h, 145.6, 187, 213, 3),
            # RIGHT column — editor-marked allergy zone top-left = (963.5, 691.6).
            # All right comments share com_x = 963.5; com_y derived per row from
            # the same pill_bottom→zone_top gap of 5.8pt (so first baseline sits
            # 12.8pt below pill bottom, matching the left column).
            ("allergy",              1095.9, 720.4, pill_w, pill_h, 961.5, 685, 213, 3),
            ("bone_and_joint",       1095.9, 622.0, pill_w, pill_h, 961.5, 586, 213, 3),
            ("posture",              1095.9, 522.4, pill_w, pill_h, 961.5, 487, 213, 3),
            ("gait_and_coordination", 1095.9, 431.6, pill_w, pill_h, 961.5, 396, 213, 3),
            ("puberty",              1095.9, 331.9, pill_w, pill_h, 961.5, 296, 213, 3),
        ]
        for key, pill_x, pill_cy, w, h, com_x, com_y, com_max_w, com_lines in rows:
            raw = (general.get(key) or {}).get("value")
            category = general_exam_normalize(key, raw) or GENERAL_DEFAULTS.get(key, "")
            if not category:
                continue
            pill_y = pill_cy - h / 2
            svg_name = GENERAL_PILL_SVGS.get(key, {}).get(category)
            if svg_name:
                self._draw_overlay_svg(c, os.path.join("Page 8", svg_name), pill_x, pill_y, w, h)
                label = os.path.splitext(os.path.basename(svg_name))[0].replace("-", " ")
                self._draw_pill_label(c, label, pill_x, pill_cy, w)
            comment = (ref.get(key, {}) or {}).get(category, "")
            if comment:
                self._draw_wrapped(
                    c, comment, x=com_x, y=com_y, max_width=com_max_w,
                    font_name=self.font("body"), font_size=8.8, leading=11,
                    color=TEXT_WHITE, max_lines=com_lines,
                )

        # Doctor Visit Recommended capsule (YES/NO) — label is baked into the art.
        dv = doctor_visit_label_from_text((general.get("doctor_visit") or {}).get("value"))
        self._draw_pill_capsule(c, DOCTOR_VISIT_SVG[dv], 966.0, 184.4)

        # Clara ID strips (left + right pages of the spread)
        self._draw_clara_footer(c, clara_id)


# ---------- CLI / module entry ----------

def generate_boy_report(json_path, backgrounds_root, output_path, fonts_folder=None,
                        xlsx_path=None, student_index=0, is_girl=False):
    """Top-level entry: load JSON, render the report (boy or girl page art)."""
    data = load_report_data(json_path, student_index)

    print(f"Patient: {data.get('student', {}).get('name', 'Unknown')}")
    print(f"DOB: {data.get('student', {}).get('dob', 'Unknown')}")
    print(f"Clara ID: {data.get('student', {}).get('clara_id', 'Unknown')}")
    print(f"Gender: {'Girl' if is_girl else 'Boy'}")
    print(f"BMI: {data.get('measurements', {}).get('bmi', '?')} -> {data.get('bmi_category', '?')}")

    gen = ClaraBoyReportGenerator(backgrounds_root, fonts_folder, xlsx_path=xlsx_path, is_girl=is_girl)
    gen.generate(data, output_path)
    print(f"Wrote {output_path}")



def generate_complete_health_report(json_path, backgrounds_folder, output_path, fonts_folder=None):
    """Compatibility entry point for pdf_service.py.

    Boy and girl share one asset folder (`backgrounds/Images`); only the page
    background differs (PageN.svg vs PageN_Girl.svg). A FEMALE student renders
    the girl backgrounds, everything else renders the boy/default backgrounds.
    """
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    entry = raw.get("data", raw)
    gender = (entry.get("student") or {}).get("gender", "").upper()
    is_girl = gender == "FEMALE"
    backgrounds_root = os.path.join(backgrounds_folder, "Images")
    xlsx_path = os.path.join(backgrounds_folder, "clara parameter updated final .xlsx")
    generate_boy_report(
        json_path=json_path,
        backgrounds_root=backgrounds_root,
        output_path=output_path,
        fonts_folder=fonts_folder,
        xlsx_path=xlsx_path,
        is_girl=is_girl,
    )

if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(here)

    generate_boy_report(
        json_path=os.path.join(project_root, "prod_sample.json"),
        backgrounds_root=os.path.join(project_root, "backgrounds", "Images"),
        output_path=os.path.join(project_root, "boy_report.pdf"),
        fonts_folder=os.path.join(project_root, "fonts"),
        xlsx_path=os.path.join(project_root, "backgrounds", "clara parameter updated final .xlsx"),
    )
